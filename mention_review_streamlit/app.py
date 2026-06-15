import csv
import io
import json
import logging
import os
import random
import re
import time
import zipfile
from urllib.parse import urlparse

import boto3
import requests
import streamlit as st
from bs4 import BeautifulSoup
from lxml import etree

from openpyxl import Workbook, load_workbook

# ── Logging ──────────────────────────────────────────────────────────────────
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="GoDaddy Mention Extractor",
    page_icon="🔍",
    layout="centered",
)

# ── Constants ─────────────────────────────────────────────────────────────────
FETCH_TIMEOUT = 20
HIT_SENTENCE_COL = "Hit Sentence"
DOMAIN_MIN_DELAY = 5.0

MELTWATER_DOMAINS = {
    "transition.meltwater.com",
    "app.meltwater.com",
    "meltwater.com",
}
MELTWATER_LOGIN_URL = "https://app.meltwater.com/"
MELTWATER_API_BASE = "https://api.meltwater.com/v3"

BEDROCK_REGION = st.secrets.get("BEDROCK_REGION", os.environ.get("BEDROCK_REGION", "us-east-2"))
BEDROCK_MODEL_ID = st.secrets.get(
    "BEDROCK_MODEL_ID",
    os.environ.get("BEDROCK_MODEL_ID", "us.anthropic.claude-sonnet-4-20250514-v1:0"),
)

EXCLUDED_TABS = {"Legend", "Hist_Data", "Hist_Calc", "Hist_Stage"}

PRESS_RELEASE_DOMAINS = {
    "prnewswire.com",
    "businesswire.com",
    "globenewswire.com",
    "prweb.com",
    "newswire.com",
    "einpresswire.com",
    "accesswire.com",
    "prlog.org",
    "send2press.com",
    "24-7pressrelease.com",
    "newswire.ca",
    "marketwired.com",
    "marketwire.com",
    "presswire.com",
    "prfire.co.uk",
    "cisionone.com",
    "cision.com",
}

USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36 Edg/122.0.0.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_4) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.4 Safari/605.1.15",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
]

# ── Helpers ───────────────────────────────────────────────────────────────────

def make_headers(url: str) -> dict:
    ua = random.choice(USER_AGENTS)
    return {
        "User-Agent": ua,
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
        "Accept-Encoding": "gzip, deflate",
        "Referer": "https://www.google.com/",
        "DNT": "1",
        "Connection": "keep-alive",
        "Upgrade-Insecure-Requests": "1",
        "Sec-Fetch-Dest": "document",
        "Sec-Fetch-Mode": "navigate",
        "Sec-Fetch-Site": "cross-site",
        "Sec-Fetch-User": "?1",
        "Cache-Control": "max-age=0",
    }


def split_sentences(text: str) -> list[str]:
    text = re.sub(r"\s+", " ", text).strip()
    text = re.sub(r"\b(Mr|Mrs|Ms|Dr|Prof|Sr|Jr|vs|etc|approx|Inc|Ltd|Corp|Co)\.", r"\1<DOT>", text)
    parts = re.split(r"(?<=[.!?])\s+(?=[A-Z\"'\u2018\u2019\u201c\u201d\(\[])", text)
    return [p.replace("<DOT>", ".").strip() for p in parts if p.strip()]


def extract_article_text(html: str) -> str:
    soup = BeautifulSoup(html, "lxml")
    for tag in soup(["script", "style", "nav", "header", "footer", "aside",
                     "noscript", "iframe", "form"]):
        tag.decompose()
    for selector in [
        "article", "main", '[role="main"]', ".post-content", ".entry-content",
        ".article-body", ".story-body", ".article__body", ".content-body",
        "#article-body", ".post-body",
    ]:
        container = soup.select_one(selector)
        if container:
            text = container.get_text(separator=" ", strip=True)
            if len(text) > 200:
                return text
    body = soup.find("body")
    return (body or soup).get_text(separator=" ", strip=True)


def extract_anchor_phrases(hit_sentence: str) -> list[str]:
    cleaned = re.sub(r"^[\s.…]+|[\s.…]+$", "", hit_sentence).strip()
    parts = re.split(r"\.{2,}|…", cleaned)
    return [p.strip() for p in parts if len(p.strip().split()) >= 4]


def find_best_context(sentences: list[str], anchor_phrases: list[str], keyword: str = "GoDaddy") -> str | None:
    keyword_lower = keyword.lower()
    best_idx, best_score = None, 0

    for i, sentence in enumerate(sentences):
        sentence_lower = sentence.lower()
        if keyword_lower not in sentence_lower:
            continue
        for phrase in anchor_phrases:
            phrase_words = set(re.findall(r"\w+", phrase.lower()))
            sent_words = set(re.findall(r"\w+", sentence_lower))
            overlap = len(phrase_words & sent_words)
            if overlap > best_score:
                best_score = overlap
                best_idx = i

    if best_idx is None:
        for i, sentence in enumerate(sentences):
            if keyword_lower in sentence.lower():
                best_idx = i
                break

    if best_idx is None:
        return None

    before = ""
    for j in range(best_idx - 1, -1, -1):
        if len(sentences[j].strip()) > 20:
            before = sentences[j]
            break

    mention = sentences[best_idx]

    after = ""
    for j in range(best_idx + 1, len(sentences)):
        if len(sentences[j].strip()) > 20:
            after = sentences[j]
            break

    result = ""
    if before:
        result += f"Before: {before}\n"
    result += f"Mention: {mention}\n"
    if after:
        result += f"After: {after}"
    return result.strip()


def fetch_with_retry(url: str, retries: int = 3) -> requests.Response | None:
    session = requests.Session()
    for attempt in range(retries):
        try:
            resp = session.get(url, headers=make_headers(url), timeout=FETCH_TIMEOUT, allow_redirects=True)
            if resp.status_code == 200:
                return resp
            if resp.status_code in (403, 429, 503) and attempt < retries - 1:
                time.sleep((attempt + 1) * 2 + random.uniform(0.5, 1.5))
                continue
            resp.raise_for_status()
        except requests.exceptions.RequestException:
            if attempt < retries - 1:
                time.sleep((attempt + 1) * 2)
                continue
            raise
    return None


def fetch_via_curl_cffi(url: str) -> str | None:
    try:
        from curl_cffi import requests as cffi_requests
        resp = cffi_requests.get(url, impersonate="chrome120", timeout=FETCH_TIMEOUT, allow_redirects=True)
        if resp.status_code == 200:
            return extract_article_text(resp.text)
        return None
    except Exception as exc:
        logger.warning("curl_cffi failed for %s: %s", url, exc)
        return None


def fetch_via_playwright(url: str) -> str | None:
    try:
        from playwright.sync_api import sync_playwright
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            context = browser.new_context(
                user_agent=random.choice(USER_AGENTS),
                locale="en-US",
                viewport={"width": 1280, "height": 800},
            )
            page = context.new_page()
            page.goto(url, wait_until="domcontentloaded", timeout=30000)
            page.wait_for_timeout(3000)
            if "msn.com" in urlparse(url).netloc:
                try:
                    btn = page.get_by_text("Continue Reading", exact=False).first
                    if btn:
                        btn.click()
                        page.wait_for_timeout(1500)
                except Exception:
                    pass
            html = page.content()
            browser.close()
            return extract_article_text(html) if html else None
    except Exception as exc:
        logger.warning("Playwright failed for %s: %s", url, exc)
        return None


def is_meltwater_url(url: str) -> bool:
    """Return True if this URL belongs to the Meltwater platform."""
    return urlparse(url).netloc in MELTWATER_DOMAINS


def _apply_cookie_string_to_session(session: requests.Session, cookies_str: str) -> None:
    """Parse a browser-copied cookie string and add each cookie to a requests Session."""
    for chunk in cookies_str.split(";"):
        chunk = chunk.strip()
        if "=" in chunk:
            name, value = chunk.split("=", 1)
            session.cookies.set(name.strip(), value.strip(), domain=".meltwater.com")


def fetch_meltwater_with_cookies(url: str, cookies_str: str) -> str | None:
    """Fetch a Meltwater URL using a browser-copied session cookie string."""
    try:
        session = requests.Session()
        _apply_cookie_string_to_session(session, cookies_str)
        resp = session.get(url, headers=make_headers(url), timeout=FETCH_TIMEOUT, allow_redirects=True)
        if resp.status_code == 200:
            text = extract_article_text(resp.text)
            # Detect if we still landed on the login/paywall page
            if any(phrase in text.lower() for phrase in ("sign in to meltwater", "log in to meltwater", "mention not found")):
                logger.warning("Meltwater cookie auth appears expired for %s", url)
                return None
            return text
        logger.warning("Meltwater cookie fetch returned HTTP %s for %s", resp.status_code, url)
        return None
    except Exception as exc:
        logger.warning("Meltwater cookie fetch failed for %s: %s", url, exc)
        return None


def extract_meltwater_doc_id(url: str) -> str | None:
    """Try to extract a document identifier from a Meltwater paywall URL."""
    parsed = urlparse(url)
    # transition.meltwater.com/paywall/redirect/<DOC_ID>?keywords=...&cid=...
    parts = parsed.path.strip("/").split("/")
    if "paywall" in parts and "redirect" in parts:
        idx = parts.index("redirect")
        if idx + 1 < len(parts):
            return parts[idx + 1]
    return None


def fetch_meltwater_via_api(url: str, api_token: str, hit_sentence: str = "") -> str | None:
    """
    Use the Meltwater REST API to search for a document matching the given URL.
    Falls back to keyword-based search using the hit_sentence if direct lookup fails.
    Returns extracted article text or None.
    """
    headers = {
        "Accept": "application/json",
        "apikey": api_token,
    }

    # Step 1: List saved searches to find one we can query against
    try:
        resp = requests.get(
            f"{MELTWATER_API_BASE}/searches",
            headers=headers,
            timeout=FETCH_TIMEOUT,
        )
        if resp.status_code == 401:
            logger.warning("Meltwater API: invalid token (401)")
            return None
        if resp.status_code != 200:
            logger.warning("Meltwater API searches returned HTTP %s", resp.status_code)
            return None

        searches = resp.json().get("searches", [])
        if not searches:
            logger.warning("Meltwater API: no saved searches found in account")
            return None

        # Step 2: Use the search endpoint to look for documents matching the URL
        # Try each saved search until we find the document
        doc_id = extract_meltwater_doc_id(url)

        for search in searches[:5]:  # Try up to 5 saved searches
            search_id = search.get("id")
            if not search_id:
                continue

            # Search for documents from the last 90 days
            try:
                search_resp = requests.get(
                    f"{MELTWATER_API_BASE}/searches/{search_id}/documents",
                    headers=headers,
                    params={"page_size": 10},
                    timeout=FETCH_TIMEOUT,
                )
                if search_resp.status_code != 200:
                    continue

                documents = search_resp.json().get("documents", [])
                for doc in documents:
                    doc_url = doc.get("url", "")
                    doc_content = doc.get("content", "") or doc.get("document_content", "")
                    doc_title = doc.get("title", "")
                    doc_summary = doc.get("summary", "") or doc.get("ingress", "")

                    # Match by URL fragment or document ID
                    if doc_url and (doc_url in url or url in doc_url):
                        text = " ".join(filter(None, [doc_title, doc_content, doc_summary]))
                        if text.strip():
                            return text.strip()
                    if doc_id and doc.get("id", "") == doc_id:
                        text = " ".join(filter(None, [doc_title, doc_content, doc_summary]))
                        if text.strip():
                            return text.strip()
            except Exception as exc:
                logger.debug("Meltwater API search %s failed: %s", search_id, exc)
                continue

    except Exception as exc:
        logger.warning("Meltwater API fetch failed: %s", exc)

    return None


def fetch_meltwater_login_and_get_cookies(email: str, password: str) -> dict | None:
    """
    Use Playwright to log in to Meltwater once and return a dict of cookies
    that can be reused for subsequent requests in this session.
    Returns None if login failed.
    """
    try:
        from playwright.sync_api import sync_playwright
        cookies_dict: dict | None = None
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            context = browser.new_context(
                user_agent=random.choice(USER_AGENTS),
                locale="en-US",
                viewport={"width": 1280, "height": 800},
            )
            page = context.new_page()

            # Navigate to the login page
            page.goto(MELTWATER_LOGIN_URL, wait_until="domcontentloaded", timeout=30000)
            page.wait_for_timeout(2000)

            # Fill email — try common selectors
            for selector in ['input[type="email"]', 'input[name="email"]', 'input[id*="email"]']:
                try:
                    page.fill(selector, email, timeout=3000)
                    break
                except Exception:
                    continue

            # Some flows show password on next screen after pressing Enter/Next
            try:
                page.get_by_role("button", name=re.compile(r"next|continue", re.IGNORECASE)).first.click(timeout=3000)
                page.wait_for_timeout(1500)
            except Exception:
                pass

            # Fill password
            for selector in ['input[type="password"]', 'input[name="password"]', 'input[id*="password"]']:
                try:
                    page.fill(selector, password, timeout=3000)
                    break
                except Exception:
                    continue

            # Submit
            try:
                page.get_by_role("button", name=re.compile(r"sign in|log in|login|submit", re.IGNORECASE)).first.click(timeout=3000)
            except Exception:
                page.keyboard.press("Enter")

            # Wait for post-login navigation
            page.wait_for_timeout(6000)

            # Check if we're past the login screen
            current_url = page.url
            if "login" in current_url.lower() or "signin" in current_url.lower():
                logger.warning("Meltwater login may have failed — still on login page: %s", current_url)
                browser.close()
                return None

            # Harvest all cookies
            raw_cookies = context.cookies()
            cookies_dict = {c["name"]: c["value"] for c in raw_cookies}
            browser.close()

        return cookies_dict if cookies_dict else None
    except Exception as exc:
        logger.warning("Meltwater Playwright login failed: %s", exc)
        return None


def fetch_meltwater_with_playwright_cookies(url: str, cookies_dict: dict) -> str | None:
    """
    Fetch a Meltwater URL using a pre-authenticated cookie dict (from a previous login).
    Uses a new Playwright context seeded with those cookies so JS-rendered content works.
    """
    try:
        from playwright.sync_api import sync_playwright
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            context = browser.new_context(
                user_agent=random.choice(USER_AGENTS),
                locale="en-US",
                viewport={"width": 1280, "height": 800},
            )
            # Seed cookies
            mw_cookies = [
                {"name": k, "value": v, "domain": ".meltwater.com", "path": "/"}
                for k, v in cookies_dict.items()
            ]
            context.add_cookies(mw_cookies)

            page = context.new_page()
            page.goto(url, wait_until="domcontentloaded", timeout=30000)
            page.wait_for_timeout(3000)
            html = page.content()
            browser.close()
            if not html:
                return None
            text = extract_article_text(html)
            # Detect paywall / login redirect
            if any(phrase in text.lower() for phrase in ("sign in to meltwater", "log in to meltwater")):
                logger.warning("Meltwater session expired mid-batch for %s", url)
                return None
            return text
    except Exception as exc:
        logger.warning("Meltwater cookie-seeded Playwright fetch failed for %s: %s", url, exc)
        return None


def fetch_and_extract(
    url: str,
    hit_sentence: str = "",
    mw_cookies_str: str = "",
    mw_playwright_cookies: dict | None = None,
    mw_api_token: str = "",
) -> str:
    if not url.startswith("http://") and not url.startswith("https://"):
        return "ERROR: Not a web URL (internal document ID)"

    article_text: str | None = None

    # ── Meltwater-specific authenticated fetch ────────────────────────────────
    if is_meltwater_url(url):
        # Try API token first (most reliable)
        if mw_api_token.strip():
            article_text = fetch_meltwater_via_api(url, mw_api_token.strip(), hit_sentence)
        # Then Playwright cookies
        if article_text is None and mw_playwright_cookies:
            article_text = fetch_meltwater_with_playwright_cookies(url, mw_playwright_cookies)
        # Then pasted cookie string
        if article_text is None and mw_cookies_str.strip():
            article_text = fetch_meltwater_with_cookies(url, mw_cookies_str)
        # ── Hit Sentence fallback for Meltwater URLs ─────────────────────────
        if article_text is None and hit_sentence.strip():
            logger.info("Meltwater auth failed for %s — falling back to Hit Sentence", url)
            # Use the Hit Sentence directly as the mention context
            cleaned = re.sub(r"^[\s.…]+|[\s.…]+$", "", hit_sentence).strip()
            if cleaned and "godaddy" in cleaned.lower():
                return f"Mention: {cleaned}\n[Source: Hit Sentence — Meltwater URL required login]"
            elif cleaned:
                # Hit Sentence exists but doesn't contain "GoDaddy" — still
                # return it so Claude can classify it rather than erroring out
                return f"Mention: {cleaned}\n[Source: Hit Sentence — Meltwater URL required login]"
        if article_text is None:
            if not mw_cookies_str.strip() and not mw_playwright_cookies and not mw_api_token.strip():
                if hit_sentence.strip():
                    cleaned = re.sub(r"^[\s.…]+|[\s.…]+$", "", hit_sentence).strip()
                    return f"Mention: {cleaned}\n[Source: Hit Sentence — no Meltwater auth provided]"
                return (
                    "ERROR: Meltwater URL requires authentication — "
                    "add credentials in the Meltwater Auth section above"
                )
            return "ERROR: Meltwater auth failed or session expired — check credentials"

    # ── Standard fetch pipeline ───────────────────────────────────────────────
    if article_text is None:
        article_text = fetch_via_curl_cffi(url)
    if not article_text:
        article_text = fetch_via_playwright(url)
    if not article_text:
        try:
            resp = fetch_with_retry(url)
            if resp is None:
                return "ERROR: Failed after retries"
            article_text = extract_article_text(resp.text)
        except requests.exceptions.Timeout:
            return "ERROR: Request timed out"
        except requests.exceptions.TooManyRedirects:
            return "ERROR: Too many redirects"
        except requests.exceptions.RequestException as exc:
            return f"ERROR: {exc}"

    sentences = split_sentences(article_text)
    anchor_phrases = extract_anchor_phrases(hit_sentence) if hit_sentence else []
    context = find_best_context(sentences, anchor_phrases)
    return context if context is not None else "ERROR: Mention not found in fetched page"


def call_llm(
    user_message: str,
    *,
    system_prompt: str = "",
    max_tokens: int = 256,
    llm_config: dict,
) -> str:
    """
    Unified LLM call. Routes to AWS Bedrock (Claude) or OpenAI based on
    llm_config['provider']. Returns the response text (stripped).

    llm_config keys:
        provider:           "bedrock" (default) or "openai"
        bedrock_client:     optional pre-built boto3 client (bedrock-runtime)
        bedrock_model_id:   defaults to module-level BEDROCK_MODEL_ID
        openai_api_key:     required when provider == "openai"
        openai_model_id:    required when provider == "openai"
    """
    provider = llm_config.get("provider", "bedrock")

    if provider == "openai":
        api_key = llm_config.get("openai_api_key", "").strip()
        model_id = llm_config.get("openai_model_id", "").strip()
        if not api_key:
            raise ValueError("OpenAI API key is required")
        if not model_id:
            raise ValueError("OpenAI model ID is required")
        try:
            from openai import OpenAI
        except ImportError as exc:
            raise RuntimeError(
                "openai package is not installed; add `openai` to requirements.txt"
            ) from exc

        client = OpenAI(api_key=api_key)
        messages = []
        if system_prompt:
            messages.append({"role": "system", "content": system_prompt})
        messages.append({"role": "user", "content": user_message})

        # Newer OpenAI models (gpt-5.x, o-series) require max_completion_tokens;
        # older models still use max_tokens. Try newer first, fall back if rejected.
        try:
            resp = client.chat.completions.create(
                model=model_id,
                messages=messages,
                max_completion_tokens=max_tokens,
            )
        except Exception as exc:
            if "max_completion_tokens" in str(exc) or "max_tokens" in str(exc):
                resp = client.chat.completions.create(
                    model=model_id,
                    messages=messages,
                    max_tokens=max_tokens,
                )
            else:
                raise
        return (resp.choices[0].message.content or "").strip()

    # Bedrock Claude (default)
    bedrock_client = llm_config.get("bedrock_client")
    if bedrock_client is None:
        bedrock_client = boto3.client("bedrock-runtime", region_name=BEDROCK_REGION)
    model_id = llm_config.get("bedrock_model_id") or BEDROCK_MODEL_ID

    body = {
        "anthropic_version": "bedrock-2023-05-31",
        "max_tokens": max_tokens,
        "messages": [{"role": "user", "content": user_message}],
    }
    if system_prompt:
        body["system"] = system_prompt

    resp = bedrock_client.invoke_model(
        modelId=model_id,
        contentType="application/json",
        accept="application/json",
        body=json.dumps(body),
    )
    return json.loads(resp["body"].read())["content"][0]["text"].strip()


def translate_to_english(text: str, llm_config: dict) -> str:
    try:
        return call_llm(
            user_message=(
                "Translate the following text to English. "
                "Preserve the Before:/Mention:/After: labels exactly as-is. "
                "Return only the translated text, nothing else.\n\n"
                f"{text}"
            ),
            max_tokens=1024,
            llm_config=llm_config,
        )
    except Exception as exc:
        logger.warning("Translation failed: %s", exc)
        return text


# ── Classification ────────────────────────────────────────────────────────────

CLASSIFICATION_SYSTEM_PROMPT = """\
You are classifying a news article's mention of "GoDaddy" based on the evidence text provided.

The evidence was extracted from the article body around where "GoDaddy" appears. It typically \
contains the sentence before the mention, the mention sentence, and the sentence after.

Classify into exactly ONE category:

A - GoDaddy Products: The evidence explicitly mentions a specific GoDaddy product, service, \
platform, tool, plan, feature, pricing tier, or product use case. \
Examples: "registered through GoDaddy", "hosted on GoDaddy", "GoDaddy Economy", \
"GoDaddy Website Builder", "GoDaddy-issued certificate", domain sold at auction by GoDaddy, \
GoDaddy as registrar, host, DNS, WHOIS, SSL, email, payments, or security provider, \
"GoDaddy website team", "GoDaddy Secret Manager".

B - GoDaddy Brand/Company: GoDaddy is mentioned only at the company or brand level with no \
concrete offering or service use case, and the evidence is not about research/data. \
Examples: employee/executive mention, stock/investor/earnings coverage, GoDaddy named as \
competitor, sponsorships, partnerships, general company references, GoDaddy quote not \
tied to research output.

C - GoDaddy Research/Data: The evidence explicitly refers to GoDaddy Small Business Research Lab, \
GoDaddy research, GoDaddy data, GoDaddy survey findings, GoDaddy reports, GoDaddy indices, \
Venture Forward, or other research/data output from the Small Business Research Lab.

D - No GoDaddy mention: The evidence contains no explicit mention of "GoDaddy" at all.

E - Error: The evidence could not be read (blank, error, inaccessible).

F - Unclear: The evidence is readable but too thin, mixed, incomplete, or contradictory \
to classify confidently.

G - Press Release: The source article is a press release or wire-service distribution. \
Key signals: content begins with or contains "FOR IMMEDIATE RELEASE"; a press-release \
dateline format (e.g. "CITY, State, Month Day, Year —"); boilerplate sections like \
"About [Company]", "Media Contact:", or a "###" end marker; highly promotional \
third-person copy announcing a product launch, partnership, or executive quote in a \
way typical of corporate press releases. Classify as G regardless of whether GoDaddy \
is also a product/brand/research mention — press releases are discarded entirely.

H - Commercial/Advertisement: The mention is about or directly references a GoDaddy \
TV commercial, advertisement, ad campaign, or promotional spot. Key signals: article \
reviews, recaps, or discusses a GoDaddy ad or commercial (e.g. Super Bowl spot, \
streaming ad); evidence references GoDaddy advertising talent or a spokesperson in \
an advertising context (e.g. "Walton Goggins", "Goggle Glasses", a celebrity \
appearing in a GoDaddy ad); coverage of a GoDaddy ad campaign, ad creative, or \
marketing stunt rather than the company or its products. Classify as H regardless \
of other signals — commercial mentions are discarded.

Strict decision rules:
- Check for press release signals FIRST. If the source is clearly a press release, \
classify as G regardless of A/B/C signals.
- Check for commercial/ad signals SECOND. If the mention is about a GoDaddy commercial \
or ad campaign, classify as H regardless of A/B/C signals.
- If both A and B signals appear, choose A.
- If both B and C signals appear and the mention is about research/survey/report/index/data, choose C.
- Do NOT use A just because the article topic is domains, hosting, or websites. \
GoDaddy itself must be the provider, tool, service, or platform in the evidence sentence.
- Do NOT use B when the evidence is about GoDaddy research, survey, report, index, or data. Use C.
- Do NOT use D if "GoDaddy" appears anywhere in the evidence text.
- If the article is mainly about another company or topic, but the evidence says the \
domain/hosting/certificate/platform is from GoDaddy, that is still A.
- For stock, investor, earnings, or company-profile stories: if no specific GoDaddy offering \
is named in the evidence, use B.
- For non-English text, "GoDaddy" usually still appears in Latin letters.

Respond with ONLY valid JSON. No markdown fences, no explanation, no preamble:
{"classification": "X", "justification": "one short sentence naming the specific evidence", "confidence": NN}

Justification rules:
- Write exactly one short sentence.
- Name the specific evidence that led to the classification.
- Good: "Evidence says the domain was registered through GoDaddy."
- Good: "Evidence cites GoDaddy Small Business Research Lab data on microbusiness trends."
- Good: "Evidence lists GoDaddy stock as an S&P 500 mover but no product is discussed."
- Good: "Content contains 'FOR IMMEDIATE RELEASE' and a press-release dateline, indicating a wire-service distribution."
- Bad: "GoDaddy is mentioned." (too vague)

Confidence scoring:
- 95-100: evidence is direct and explicit
- 85-94: evidence is clear but slightly indirect or brief
- 70-84: evidence supports the label but some ambiguity exists
- 50-69: evidence is limited or mixed
- Below 50: substantial uncertainty
"""


def classify_mention(evidence: str, llm_config: dict, url: str = "") -> dict:
    """
    Send evidence text to the configured LLM and return classification dict.
    Returns {"classification": "X", "justification": "...", "confidence": N}.
    """
    # Pre-check: known press release wire service domains → auto-classify as G
    if url:
        netloc = urlparse(url).netloc.lower().lstrip("www.")
        if any(netloc == d or netloc.endswith("." + d) for d in PRESS_RELEASE_DOMAINS):
            return {
                "classification": "G",
                "justification": f"Article URL is from a known press wire service ({netloc}).",
                "confidence": 98,
            }

    default_error = {
        "classification": "E",
        "justification": "LLM classification call failed.",
        "confidence": 90,
    }
    try:
        raw = call_llm(
            user_message=f"Classify this evidence:\n\n{evidence}",
            system_prompt=CLASSIFICATION_SYSTEM_PROMPT,
            max_tokens=256,
            llm_config=llm_config,
        )
        # Strip markdown fences if present
        raw = re.sub(r"^```(?:json)?\s*", "", raw)
        raw = re.sub(r"\s*```$", "", raw)
        result = json.loads(raw)
        # Validate
        if result.get("classification") not in ("A", "B", "C", "D", "E", "F", "G", "H"):
            result["classification"] = "F"
        result["confidence"] = int(result.get("confidence", 50))
        return result
    except Exception as exc:
        logger.warning("Classification failed: %s", exc)
        return default_error


def normalize_column(col: str, headers: list[str]) -> str | None:
    col = col.strip()
    for h in headers:
        if h.strip().lower() == col.lower():
            return h
    if re.fullmatch(r"[A-Za-z]{1,2}", col):
        idx = 0
        for ch in col.upper():
            idx = idx * 26 + (ord(ch) - ord("A") + 1)
        idx -= 1
        if 0 <= idx < len(headers):
            return headers[idx]
    return None


def read_uploaded_table(uploaded_file) -> tuple[list[str], list[dict]]:
    """
    Read an uploaded CSV or XLSX file into (headers, rows).
    `rows` is a list of dicts keyed by header. Returns ([], []) on failure.
    """
    name = (getattr(uploaded_file, "name", "") or "").lower()
    raw = uploaded_file.read()

    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        try:
            wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
        except Exception as exc:
            logger.warning("Failed to open xlsx: %s", exc)
            return [], []
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            wb.close()
            return [], []
        headers = [(str(h).strip() if h is not None else "") for h in header_row]
        rows: list[dict] = []
        for r in rows_iter:
            if r is None:
                continue
            if all(v is None or (isinstance(v, str) and not v.strip()) for v in r):
                continue
            row_dict = {}
            for i, h in enumerate(headers):
                if not h:
                    continue
                val = r[i] if i < len(r) else None
                row_dict[h] = "" if val is None else str(val)
            rows.append(row_dict)
        wb.close()
        return headers, rows

    # CSV path
    content: str | None = None
    for enc in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            content = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if content is None:
        return [], []
    reader = csv.DictReader(io.StringIO(content))
    headers = list(reader.fieldnames or [])
    rows = list(reader)
    return headers, rows


def rows_to_xlsx_bytes(headers: list[str], rows: list[dict]) -> bytes:
    """
    Write rows to an .xlsx byte buffer. Uses openpyxl defaults so every
    row keeps Excel's default row height (no wrap_text, no custom dimensions),
    which is what the user wants regardless of multi-line cell content.
    """
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


def interleave_by_domain(rows: list[dict], url_col: str) -> list[dict]:
    domain_groups: dict[str, list] = {}
    for row in rows:
        domain = urlparse(row.get(url_col, "").strip()).netloc or ""
        domain_groups.setdefault(domain, []).append(row)
    interleaved = []
    while any(domain_groups.values()):
        for d in list(domain_groups.keys()):
            if domain_groups[d]:
                interleaved.append(domain_groups[d].pop(0))
            if not domain_groups[d]:
                del domain_groups[d]
    return interleaved


def get_xlsx_sheet_names(raw_bytes: bytes) -> list[str]:
    """Return selectable sheet names from an XLSX, excluding Hist_* tabs."""
    try:
        wb = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
        names = [n for n in wb.sheetnames if n not in EXCLUDED_TABS]
        wb.close()
        return names
    except Exception as exc:
        logger.warning("Failed to read sheet names: %s", exc)
        return []


def _build_red_xf_set(raw_bytes: bytes) -> set[int]:
    """
    Parse the xlsx styles XML and return the set of XF (cell-style) indices
    whose fill is solid red (#FF0000).  Rows containing any cell with one of
    these styles should be discarded (Legend: Red = "Removed").
    """
    XNS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    sns = {"x": XNS}

    with zipfile.ZipFile(io.BytesIO(raw_bytes)) as zf:
        styles_xml = zf.read("xl/styles.xml")

    sroot = etree.fromstring(styles_xml)

    red_fill_ids: set[int] = set()
    for i, fill in enumerate(sroot.findall(".//x:fills/x:fill", sns)):
        pf = fill.find("x:patternFill", sns)
        if pf is None:
            continue
        fg = pf.find("x:fgColor", sns)
        if fg is None:
            continue
        rgb = fg.get("rgb", "")
        if rgb == "FFFF0000":
            red_fill_ids.add(i)

    red_xfs: set[int] = set()
    for idx, xf in enumerate(sroot.findall(".//x:cellXfs/x:xf", sns)):
        fill_id = int(xf.get("fillId", 0))
        if fill_id in red_fill_ids:
            red_xfs.add(idx)

    return red_xfs


def _get_red_row_nums(raw_bytes: bytes, sheet_name: str) -> set[int]:
    """
    Return the set of row numbers in *sheet_name* that contain at least one
    cell with a red (#FF0000) fill.  These are the "Removed" rows per the
    workbook Legend and should be discarded.
    """
    XNS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    RELS_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
    sns = {"x": XNS}

    red_xfs = _build_red_xf_set(raw_bytes)
    if not red_xfs:
        return set()

    with zipfile.ZipFile(io.BytesIO(raw_bytes)) as zf:
        wb_xml = zf.read("xl/workbook.xml")
        wb_root = etree.fromstring(wb_xml)
        rels_xml = zf.read("xl/_rels/workbook.xml.rels")
        rels_root = etree.fromstring(rels_xml)

        rid_to_target: dict[str, str] = {}
        for rel in rels_root.findall(f"{{{RELS_NS}}}Relationship"):
            rid_to_target[rel.get("Id")] = rel.get("Target")

        sheet_rid = None
        for s in wb_root.findall(".//x:sheets/x:sheet", sns):
            if s.get("name") == sheet_name:
                ns_r = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
                sheet_rid = s.get(f"{{{ns_r}}}id")
                break

        if sheet_rid is None or sheet_rid not in rid_to_target:
            return set()

        sheet_path = "xl/" + rid_to_target[sheet_rid].lstrip("/")
        sheet_xml = zf.read(sheet_path)

    sheet_root = etree.fromstring(sheet_xml)
    red_rows: set[int] = set()
    for xml_row in sheet_root.findall(f".//{{{XNS}}}sheetData/{{{XNS}}}row"):
        row_num = int(xml_row.get("r"))
        for cell in xml_row.findall(f"{{{XNS}}}c"):
            s = int(cell.get("s", 0))
            if s in red_xfs:
                red_rows.add(row_num)
                break
    return red_rows


def get_xlsx_date_range(raw_bytes: bytes, sheet_name: str) -> tuple:
    """
    Return (min_date, max_date) found in the Date column of *sheet_name*.
    Returns (None, None) if dates can't be determined.
    """
    import datetime
    try:
        wb = load_workbook(io.BytesIO(raw_bytes), data_only=True, read_only=True)
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if header_row is None:
            wb.close()
            return None, None
        headers = [(str(h).strip() if h is not None else "") for h in header_row]
        date_idx = None
        for i, h in enumerate(headers):
            if h.lower() == "date":
                date_idx = i
                break
        if date_idx is None:
            wb.close()
            return None, None
        dates = []
        for vals in rows_iter:
            if vals is None:
                continue
            v = vals[date_idx] if date_idx < len(vals) else None
            if isinstance(v, datetime.datetime):
                dates.append(v.date())
            elif isinstance(v, datetime.date):
                dates.append(v)
        wb.close()
        if not dates:
            return None, None
        return min(dates), max(dates)
    except Exception as exc:
        logger.warning("Could not read date range: %s", exc)
        return None, None


def read_uploaded_table_filtered(
    uploaded_file,
    sheet_name: str | None = None,
    date_start=None,
    date_end=None,
) -> tuple[list[str], list[dict]]:
    """
    Read an uploaded XLSX, selecting a specific sheet and:
      1. Discarding any rows marked red (Legend: "Removed").
      2. Optionally filtering to rows whose Date falls within [date_start, date_end].
    Falls back to the unfiltered ``read_uploaded_table`` for CSVs or when
    no sheet is selected.
    """
    import datetime
 
    name = (getattr(uploaded_file, "name", "") or "").lower()
    raw = uploaded_file.read()
 
    if not (name.endswith(".xlsx") or name.endswith(".xlsm")) or sheet_name is None:
        uploaded_file.seek(0)
        return read_uploaded_table(uploaded_file)
 
    # ── Identify red rows to discard ─────────────────────────────────────────
    red_row_nums = _get_red_row_nums(raw, sheet_name)
 
    # ── Read the chosen sheet via openpyxl (read_only=False for reliability) ─
    try:
        wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=False)
    except Exception as exc:
        logger.warning("Failed to open xlsx: %s", exc)
        return [], []
 
    ws = wb[sheet_name]
    if ws.max_row is None or ws.max_row < 2 or ws.max_column is None:
        wb.close()
        return [], []
 
    max_col = ws.max_column
    headers = [(str(ws.cell(row=1, column=c).value or "").strip())
                for c in range(1, max_col + 1)]
 
    # Find Date column index for date filtering (0-based)
    date_col_idx: int | None = None
    if date_start is not None or date_end is not None:
        for i, h in enumerate(headers):
            if h.lower() == "date":
                date_col_idx = i
                break
 
    rows: list[dict] = []
    for r in range(2, ws.max_row + 1):
        # Skip red "Removed" rows
        if r in red_row_nums:
            continue
 
        # Read all cell values for this row
        vals = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
 
        # Skip entirely blank rows
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in vals):
            continue
 
        # Date range filter
        if date_col_idx is not None:
            cell_val = vals[date_col_idx]
            cell_date = None
            if isinstance(cell_val, datetime.datetime):
                cell_date = cell_val.date()
            elif isinstance(cell_val, datetime.date):
                cell_date = cell_val
            if cell_date is None:
                continue  # skip rows without a parseable date
            if date_start and cell_date < date_start:
                continue
            if date_end and cell_date > date_end:
                continue
 
        row_dict = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            val = vals[i]
            row_dict[h] = "" if val is None else str(val)
        rows.append(row_dict)
 
    wb.close()
    return headers, rows


# ── UI ────────────────────────────────────────────────────────────────────────

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=DM+Mono:wght@400;500&family=DM+Sans:wght@300;400;500;600&display=swap');

    html, body, [class*="css"] {
        font-family: 'DM Sans', sans-serif;
    }
    .stApp {
        background: #f7f6f3;
    }
    .block-container {
        padding-top: 2.5rem;
        max-width: 760px;
    }
    h1 {
        font-family: 'DM Mono', monospace !important;
        font-size: 1.4rem !important;
        font-weight: 500 !important;
        letter-spacing: -0.02em;
        color: #1a1a1a;
        border-bottom: 2px solid #1a1a1a;
        padding-bottom: 0.6rem;
        margin-bottom: 0.2rem !important;
    }
    .subtitle {
        font-size: 0.85rem;
        color: #666;
        margin-bottom: 2rem;
        font-family: 'DM Mono', monospace;
    }
    .stButton > button {
        background: #1a1a1a;
        color: #f7f6f3;
        border: none;
        border-radius: 4px;
        font-family: 'DM Mono', monospace;
        font-size: 0.85rem;
        letter-spacing: 0.04em;
        padding: 0.55rem 1.4rem;
        transition: opacity 0.15s;
        width: 100%;
    }
    .stButton > button:hover {
        opacity: 0.8;
        color: #f7f6f3;
    }
    .stTextInput > div > div > input,
    .stSelectbox > div > div {
        font-family: 'DM Mono', monospace;
        font-size: 0.85rem;
        border-radius: 4px;
        border-color: #d0cfc9;
        background: #fff;
    }
    label {
        font-size: 0.8rem !important;
        font-weight: 500 !important;
        color: #444 !important;
        text-transform: uppercase;
        letter-spacing: 0.06em;
    }
    .stat-row {
        display: flex;
        gap: 1rem;
        margin: 1rem 0;
    }
    .stat-box {
        flex: 1;
        background: #fff;
        border: 1px solid #e0dfd9;
        border-radius: 6px;
        padding: 0.9rem 1rem;
        text-align: center;
    }
    .stat-num {
        font-family: 'DM Mono', monospace;
        font-size: 1.6rem;
        font-weight: 500;
        color: #1a1a1a;
        line-height: 1;
    }
    .stat-label {
        font-size: 0.7rem;
        text-transform: uppercase;
        letter-spacing: 0.08em;
        color: #888;
        margin-top: 0.3rem;
    }
    .stat-num.errors { color: #c0392b; }
    .stDownloadButton > button {
        background: #fff;
        color: #1a1a1a;
        border: 2px solid #1a1a1a;
        border-radius: 4px;
        font-family: 'DM Mono', monospace;
        font-size: 0.85rem;
        letter-spacing: 0.04em;
        width: 100%;
        margin-top: 0.5rem;
    }
    .stDownloadButton > button:hover {
        background: #1a1a1a;
        color: #f7f6f3;
    }
    hr {
        border: none;
        border-top: 1px solid #e0dfd9;
        margin: 1.5rem 0;
    }
    .stCheckbox label {
        text-transform: none !important;
        letter-spacing: normal !important;
        font-size: 0.9rem !important;
    }
    .stProgress > div > div > div {
        background: #1a1a1a;
    }
    [data-testid="stFileUploader"] {
        background: #fff;
        border: 1px dashed #c0bfba;
        border-radius: 6px;
        padding: 0.5rem;
    }
</style>
""", unsafe_allow_html=True)

st.markdown("# GoDaddy Mention Extractor")
st.markdown('<p class="subtitle">Upload a CSV or XLSX → extract mention context → download enriched XLSX</p>', unsafe_allow_html=True)

# ── Model Selection ───────────────────────────────────────────────────────────
OPENAI_MODEL_PRESETS = [
    "gpt-5.5",
    "gpt-5.4-mini",
    "gpt-5.4-nano",
    "gpt-5.2",
    "gpt-4o",
    "gpt-4o-mini",
    "Custom…",
]

with st.expander("🤖 Model Selection", expanded=False):
    st.markdown(
        "<small style='color:#666;font-family:DM Mono,monospace'>"
        "Choose which LLM powers translation (Step 1) and classification (Step 2). "
        "Defaults to AWS Bedrock with the configured Claude model.</small>",
        unsafe_allow_html=True,
    )
    llm_provider = st.radio(
        "LLM Provider",
        ["AWS Bedrock (Claude)", "OpenAI (ChatGPT)"],
        horizontal=True,
        key="llm_provider",
    )

    openai_api_key_input = ""
    openai_model_id_input = ""

    if llm_provider == "OpenAI (ChatGPT)":
        st.markdown(
            "<small style='color:#888;font-family:DM Mono,monospace'>"
            "Get a key at <b>platform.openai.com → API keys</b>. "
            "The key is held in session memory only and never written to disk.</small>",
            unsafe_allow_html=True,
        )
        openai_api_key_input = st.text_input(
            "OpenAI API Key",
            type="password",
            key="openai_api_key",
            placeholder="sk-…",
        )
        oa_col1, oa_col2 = st.columns([1, 1])
        with oa_col1:
            openai_model_choice = st.selectbox(
                "OpenAI Model",
                OPENAI_MODEL_PRESETS,
                index=0,
                key="openai_model_choice",
            )
        with oa_col2:
            if openai_model_choice == "Custom…":
                openai_model_id_input = st.text_input(
                    "Custom Model ID",
                    value="",
                    key="openai_model_custom",
                    placeholder="e.g. gpt-5.5-2025-12-01",
                )
            else:
                openai_model_id_input = openai_model_choice
                st.markdown(
                    f"<div style='font-family:DM Mono,monospace;font-size:0.8rem;"
                    f"color:#555;padding-top:1.9rem'>Using <b>{openai_model_id_input}</b></div>",
                    unsafe_allow_html=True,
                )
    else:
        st.markdown(
            f"<small style='color:#888;font-family:DM Mono,monospace'>"
            f"Bedrock region: <b>{BEDROCK_REGION}</b> · Model: <b>{BEDROCK_MODEL_ID}</b></small>",
            unsafe_allow_html=True,
        )


def build_llm_config() -> tuple[dict, str | None]:
    """
    Build the llm_config dict from current UI / session state.
    Returns (config, error_message). error_message is None on success.
    """
    provider_choice = st.session_state.get("llm_provider", "AWS Bedrock (Claude)")
    if provider_choice == "OpenAI (ChatGPT)":
        api_key = st.session_state.get("openai_api_key", "").strip()
        choice = st.session_state.get("openai_model_choice", "")
        model_id = (
            st.session_state.get("openai_model_custom", "").strip()
            if choice == "Custom…"
            else choice
        )
        if not api_key:
            return {}, "OpenAI is selected but no API key was provided. Open the Model Selection section and paste your key."
        if not model_id:
            return {}, "OpenAI is selected but no model ID was provided."
        return {
            "provider": "openai",
            "openai_api_key": api_key,
            "openai_model_id": model_id,
        }, None

    return {
        "provider": "bedrock",
        "bedrock_model_id": BEDROCK_MODEL_ID,
    }, None


# ── Form ──────────────────────────────────────────────────────────────────────
uploaded_file = st.file_uploader("CSV or XLSX File", type=["csv", "xlsx"], label_visibility="visible")

# ── Sheet Tab Selector (XLSX only) ───────────────────────────────────────────
selected_sheet: str | None = None
filter_date_start = None
filter_date_end = None
_xlsx_raw_cache: bytes | None = None

if uploaded_file is not None and (uploaded_file.name or "").lower().endswith((".xlsx", ".xlsm")):
    _xlsx_raw_cache = uploaded_file.read()
    uploaded_file.seek(0)
    sheet_names = get_xlsx_sheet_names(_xlsx_raw_cache)
    if sheet_names:
        selected_sheet = st.selectbox(
            "Excel Sheet Tab",
            sheet_names,
            index=0,
            help="Choose which sheet to process. Rows marked red (Removed) are automatically discarded.",
            key="sheet_tab",
        )
        # Date range filter
        if selected_sheet:
            d_min, d_max = get_xlsx_date_range(_xlsx_raw_cache, selected_sheet)
            if d_min and d_max:
                date_col1, date_col2 = st.columns(2)
                with date_col1:
                    filter_date_start = st.date_input(
                        "Start Date",
                        value=d_min,
                        min_value=d_min,
                        max_value=d_max,
                        help="Include rows on or after this date",
                        key="filter_date_start",
                    )
                with date_col2:
                    filter_date_end = st.date_input(
                        "End Date",
                        value=d_max,
                        min_value=d_min,
                        max_value=d_max,
                        help="Include rows on or before this date",
                        key="filter_date_end",
                    )

col1, col2 = st.columns(2)
with col1:
    url_col_input = st.text_input("URL Column", value="URL",
                                   help="Column header name (e.g. 'URL') or letter (e.g. 'B')")
with col2:
    output_col_name = st.text_input("Output Column Name", value="GoDaddy Mentions")

col3, col4 = st.columns([2, 1])
with col3:
    delay = st.slider("Delay between requests (s)", min_value=0.5, max_value=5.0, value=1.5, step=0.5)
with col4:
    translate = st.checkbox("Translate to English", value=False,
                             help="Uses AWS Bedrock (Claude) to translate non-English mentions")

# ── Meltwater Authentication ───────────────────────────────────────────────────
with st.expander("🔐 Meltwater Authentication", expanded=False):
    st.markdown(
        "<small style='color:#666;font-family:DM Mono,monospace'>"
        "Broadcast transcript URLs on Meltwater (transition.meltwater.com) require login. "
        "Choose how to authenticate below. If no method is selected, the app will "
        "fall back to the <b>Hit Sentence</b> column for Meltwater URLs.</small>",
        unsafe_allow_html=True,
    )
    mw_auth_method = st.radio(
        "Authentication method",
        ["None (use Hit Sentence fallback)", "API Token", "Email & Password (Playwright)", "Paste Session Cookie"],
        horizontal=True,
        label_visibility="collapsed",
    )

    mw_email, mw_password, mw_cookie_str, mw_api_token = "", "", "", ""

    if mw_auth_method == "API Token":
        st.markdown(
            "<small style='color:#888;font-family:DM Mono,monospace'>"
            "Generate a token in Meltwater: <b>Account → Meltwater API → Create Token</b>. "
            "This uses the Meltwater REST API to search for document content. "
            "Requires API access on your Meltwater subscription.</small>",
            unsafe_allow_html=True,
        )
        mw_api_token = st.text_input("Meltwater API Token", type="password", key="mw_api_token")

    elif mw_auth_method == "Email & Password (Playwright)":
        st.markdown(
            "<small style='color:#888;font-family:DM Mono,monospace'>"
            "Playwright will log in once and reuse the session for all Meltwater URLs in this run. "
            "Credentials are never stored.</small>",
            unsafe_allow_html=True,
        )
        mw_col1, mw_col2 = st.columns(2)
        with mw_col1:
            mw_email = st.text_input("Meltwater Email", key="mw_email")
        with mw_col2:
            mw_password = st.text_input("Meltwater Password", type="password", key="mw_password")

    elif mw_auth_method == "Paste Session Cookie":
        st.markdown(
            "<small style='color:#888;font-family:DM Mono,monospace'>"
            "Open a Meltwater page while logged in → DevTools → Application → Cookies → "
            "copy the full cookie string and paste it here.</small>",
            unsafe_allow_html=True,
        )
        mw_cookie_str = st.text_area(
            "Cookie string (name=value; name2=value2; …)",
            height=80,
            key="mw_cookie_str",
            placeholder="__cf_bm=abc123; mw_session=xyz789; …",
        )

st.markdown("<hr>", unsafe_allow_html=True)

# ── Step 1: Extract Mentions (rerun-resilient) ───────────────────────────────
# The button press validates inputs, stores everything into session_state, and
# sets a running flag.  The actual processing loop runs off that flag so that
# if Streamlit triggers a rerun (e.g. tab hidden → WebSocket reconnect) the
# loop picks up where it left off instead of silently dying.

run_button = st.button("Extract Mentions", disabled=st.session_state.get("ext_running", False))

if run_button:
    # ── Validate inputs ──────────────────────────────────────────────────────
    if not uploaded_file:
        st.error("Please upload a CSV or XLSX file first.")
        st.stop()
    if not url_col_input.strip():
        st.error("Please specify the URL column.")
        st.stop()

    llm_config, llm_err = build_llm_config()
    if translate and llm_err:
        st.error(llm_err)
        st.stop()

    headers, rows = read_uploaded_table_filtered(
        uploaded_file, selected_sheet,
        date_start=filter_date_start, date_end=filter_date_end,
    )
    if not headers:
        st.error("Could not read the uploaded file. Make sure it's a valid CSV or XLSX with a header row.")
        st.stop()

    if not rows and selected_sheet is not None:
        st.error(f"No rows found on sheet '{selected_sheet}' after filtering. Check that the date range contains data and not all rows are marked red.")
        st.stop()

    url_col = normalize_column(url_col_input, headers)
    if url_col is None:
        st.error(f"Column '{url_col_input}' not found. Available: {', '.join(headers)}")
        st.stop()

    if not rows:
        st.error("File has no data rows.")
        st.stop()

    interleaved = interleave_by_domain(rows, url_col)
    out_headers = headers + ([output_col_name] if output_col_name not in headers else [])

    # ── Meltwater pre-login (once per run) ───────────────────────────────────
    mw_playwright_cookies: dict | None = None
    has_mw_urls = any(is_meltwater_url(r.get(url_col, "").strip()) for r in interleaved)

    if has_mw_urls and mw_auth_method == "API Token":
        if mw_api_token.strip():
            with st.spinner("🔐 Validating Meltwater API token…"):
                try:
                    test_resp = requests.get(
                        f"{MELTWATER_API_BASE}/searches",
                        headers={"Accept": "application/json", "apikey": mw_api_token.strip()},
                        timeout=FETCH_TIMEOUT,
                    )
                    if test_resp.status_code == 200:
                        count = len(test_resp.json().get("searches", []))
                        st.success(f"✓ Meltwater API token valid ({count} saved searches found)")
                    elif test_resp.status_code == 401:
                        st.error("⚠ Meltwater API token is invalid (401 Unauthorized)")
                        mw_api_token = ""
                    else:
                        st.warning(f"⚠ Meltwater API returned HTTP {test_resp.status_code} — will try anyway")
                except Exception as exc:
                    st.warning(f"⚠ Could not validate token: {exc} — will try anyway")
        else:
            st.warning("⚠ Meltwater URLs detected but no API token provided — will fall back to Hit Sentence")

    elif has_mw_urls and mw_auth_method == "Email & Password (Playwright)":
        if mw_email and mw_password:
            with st.spinner("🔐 Logging in to Meltwater…"):
                mw_playwright_cookies = fetch_meltwater_login_and_get_cookies(mw_email, mw_password)
            if mw_playwright_cookies:
                st.success(f"✓ Meltwater login succeeded ({len(mw_playwright_cookies)} cookies captured)")
            else:
                st.error("⚠ Meltwater login failed — Playwright could not authenticate. Check credentials.")
        else:
            st.warning("⚠ Meltwater URLs detected but no email/password provided — those rows will be skipped.")

    elif has_mw_urls and mw_auth_method == "None (use Hit Sentence fallback)":
        st.info("ℹ Meltwater URLs detected — will use Hit Sentence column as fallback for those rows.")

    # ── Snapshot everything into session_state so reruns can resume ───────────
    st.session_state["ext_running"] = True
    st.session_state["ext_progress"] = 0
    st.session_state["ext_ok"] = 0
    st.session_state["ext_err"] = 0
    # Store rows by index (not id()) so ordering survives reruns
    st.session_state["ext_interleaved"] = interleaved
    st.session_state["ext_original_rows"] = rows
    st.session_state["ext_headers"] = headers
    st.session_state["ext_out_headers"] = out_headers
    st.session_state["ext_url_col"] = url_col
    st.session_state["ext_output_col"] = output_col_name
    st.session_state["ext_delay"] = delay
    st.session_state["ext_translate"] = translate
    st.session_state["ext_llm_config"] = {k: v for k, v in (llm_config or {}).items() if k != "bedrock_client"}
    st.session_state["ext_mw_cookie_str"] = mw_cookie_str
    st.session_state["ext_mw_playwright_cookies"] = mw_playwright_cookies
    st.session_state["ext_mw_api_token"] = mw_api_token
    st.session_state["ext_results"] = {}          # idx → result string
    st.session_state["ext_domain_last_hit"] = {}   # domain → timestamp
    st.rerun()

# ── Resume / run extraction loop ─────────────────────────────────────────────
if st.session_state.get("ext_running"):
    interleaved = st.session_state["ext_interleaved"]
    original_rows = st.session_state["ext_original_rows"]
    ext_headers = st.session_state["ext_headers"]
    out_headers = st.session_state["ext_out_headers"]
    url_col = st.session_state["ext_url_col"]
    output_col_name_run = st.session_state["ext_output_col"]
    ext_delay = st.session_state["ext_delay"]
    ext_translate = st.session_state["ext_translate"]
    ext_llm_config = st.session_state["ext_llm_config"]
    ext_mw_cookie_str = st.session_state["ext_mw_cookie_str"]
    ext_mw_playwright_cookies = st.session_state["ext_mw_playwright_cookies"]
    ext_mw_api_token = st.session_state["ext_mw_api_token"]
    ext_results = st.session_state["ext_results"]
    domain_last_hit = st.session_state["ext_domain_last_hit"]
    start_idx = st.session_state["ext_progress"]

    total = len(interleaved)
    ok_count = st.session_state["ext_ok"]
    error_count = st.session_state["ext_err"]

    # Rebuild bedrock client if needed (can't be serialised into session_state)
    if ext_llm_config.get("provider") == "bedrock" and "bedrock_client" not in ext_llm_config:
        try:
            ext_llm_config["bedrock_client"] = boto3.client("bedrock-runtime", region_name=BEDROCK_REGION)
        except Exception as exc:
            st.error(f"Could not reconnect to AWS Bedrock: {exc}")
            st.session_state["ext_running"] = False
            st.stop()

    # Cancel button
    if st.button("⛔ Cancel Extraction", key="cancel_ext"):
        st.session_state["ext_running"] = False
        st.warning("Extraction cancelled.")
        st.rerun()

    status_text = st.empty()
    progress_bar = st.progress(start_idx / total if total else 0)
    stats_placeholder = st.empty()

    # Show stats for already-completed rows
    if start_idx > 0:
        status_text.markdown(
            f"<small style='font-family:DM Mono,monospace;color:#888'>▶ Resuming from row {start_idx+1}/{total}…</small>",
            unsafe_allow_html=True,
        )

    for i in range(start_idx, total):
        row = interleaved[i]
        url = row.get(url_col, "").strip()
        hit_sentence = row.get(HIT_SENTENCE_COL, "").strip()

        if url:
            domain = urlparse(url).netloc
            if domain in domain_last_hit:
                elapsed = time.time() - domain_last_hit[domain]
                if elapsed < DOMAIN_MIN_DELAY:
                    wait = DOMAIN_MIN_DELAY - elapsed + random.uniform(0.5, 1.5)
                    status_text.markdown(
                        f"<small style='font-family:DM Mono,monospace;color:#888'>⏳ Throttling `{domain}` — {wait:.1f}s</small>",
                        unsafe_allow_html=True,
                    )
                    time.sleep(wait)

            status_text.markdown(
                f"<small style='font-family:DM Mono,monospace;color:#555'>🔍 {url[:80]}{'…' if len(url)>80 else ''}</small>",
                unsafe_allow_html=True,
            )
            result = fetch_and_extract(
                url,
                hit_sentence,
                mw_cookies_str=ext_mw_cookie_str,
                mw_playwright_cookies=ext_mw_playwright_cookies,
                mw_api_token=ext_mw_api_token,
            )
            domain_last_hit[domain] = time.time()

            if ext_translate and result and not result.startswith("ERROR"):
                lang = row.get("Language", "").strip().lower()
                if lang and lang != "english":
                    status_text.markdown(
                        f"<small style='font-family:DM Mono,monospace;color:#888'>🌐 Translating ({lang})…</small>",
                        unsafe_allow_html=True,
                    )
                    result = translate_to_english(result, ext_llm_config)

            row[output_col_name_run] = result
            if result.startswith("ERROR"):
                error_count += 1
            else:
                ok_count += 1

            if i < total - 1:
                time.sleep(ext_delay + random.uniform(0.5, 1.0))
        else:
            row[output_col_name_run] = "NO_URL"

        # ── Persist progress so reruns can resume ────────────────────────────
        ext_results[i] = row[output_col_name_run]
        st.session_state["ext_progress"] = i + 1
        st.session_state["ext_ok"] = ok_count
        st.session_state["ext_err"] = error_count
        st.session_state["ext_results"] = ext_results
        st.session_state["ext_domain_last_hit"] = domain_last_hit

        progress_bar.progress((i + 1) / total)
        stats_placeholder.markdown(
            f"""<div class="stat-row">
                <div class="stat-box"><div class="stat-num">{i+1}/{total}</div><div class="stat-label">Processed</div></div>
                <div class="stat-box"><div class="stat-num">{ok_count}</div><div class="stat-label">Extracted</div></div>
                <div class="stat-box"><div class="stat-num errors">{error_count}</div><div class="stat-label">Errors</div></div>
            </div>""",
            unsafe_allow_html=True,
        )

    # ── Complete — build final output ────────────────────────────────────────
    # Re-apply results to original row order
    # Build a mapping: original row index → interleaved index
    interleaved_id_order = {id(r): idx for idx, r in enumerate(interleaved)}
    for orig_row in original_rows:
        il_idx = interleaved_id_order.get(id(orig_row))
        if il_idx is not None and il_idx in ext_results:
            orig_row[output_col_name_run] = ext_results[il_idx]

    ordered = original_rows
    xlsx_bytes = rows_to_xlsx_bytes(out_headers, ordered)

    status_text.markdown(
        "<small style='font-family:DM Mono,monospace;color:#1a1a1a'>✓ Done</small>",
        unsafe_allow_html=True,
    )

    st.markdown("<hr>", unsafe_allow_html=True)
    st.download_button(
        label="⬇ Download Enriched XLSX",
        data=xlsx_bytes,
        file_name="mentions_output.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    # Store enriched data in session state for Step 2
    st.session_state["enriched_rows"] = ordered
    st.session_state["enriched_headers"] = out_headers
    st.session_state["enriched_output_col"] = output_col_name_run

    # Clear running state
    st.session_state["ext_running"] = False

# ── Step 2: Classify Mentions ─────────────────────────────────────────────────
st.markdown("<hr>", unsafe_allow_html=True)
st.markdown("# Classify Mentions")
st.markdown(
    '<p class="subtitle">Classify each extracted mention into A (Product) · B (Brand) · C (Research) · D (None) · E (Error) · F (Unclear) · G (Press Release) · H (Commercial)</p>',
    unsafe_allow_html=True,
)

classify_source = st.radio(
    "Source file",
    ["Use output from Step 1 above", "Upload a pre-enriched file"],
    horizontal=True,
    label_visibility="collapsed",
    key="classify_source",
)

classify_file = None
classify_selected_sheet: str | None = None
classify_date_start = None
classify_date_end = None
if classify_source == "Upload a pre-enriched file":
    classify_file = st.file_uploader(
        "CSV or XLSX with extracted mentions",
        type=["csv", "xlsx"],
        key="classify_upload",
    )
    if classify_file is not None and (classify_file.name or "").lower().endswith((".xlsx", ".xlsm")):
        cl_raw = classify_file.read()
        classify_file.seek(0)
        cl_sheet_names = get_xlsx_sheet_names(cl_raw)
        if cl_sheet_names:
            classify_selected_sheet = st.selectbox(
                "Excel Sheet Tab",
                cl_sheet_names,
                index=0,
                help="Choose which sheet to classify. Rows marked red are automatically discarded.",
                key="classify_sheet_tab",
            )
            if classify_selected_sheet:
                cl_d_min, cl_d_max = get_xlsx_date_range(cl_raw, classify_selected_sheet)
                if cl_d_min and cl_d_max:
                    cl_date_col1, cl_date_col2 = st.columns(2)
                    with cl_date_col1:
                        classify_date_start = st.date_input(
                            "Start Date",
                            value=cl_d_min,
                            min_value=cl_d_min,
                            max_value=cl_d_max,
                            key="classify_date_start",
                        )
                    with cl_date_col2:
                        classify_date_end = st.date_input(
                            "End Date",
                            value=cl_d_max,
                            min_value=cl_d_min,
                            max_value=cl_d_max,
                            key="classify_date_end",
                        )

cl_col1, cl_col2 = st.columns(2)
with cl_col1:
    evidence_col_input = st.text_input(
        "Evidence Column (AQ)",
        value="GoDaddy Mentions",
        help="Column containing the extracted mention text (output of Step 1)",
        key="evidence_col",
    )
with cl_col2:
    classify_url_col_input = st.text_input(
        "URL Column",
        value="URL",
        help="Column with article URLs (for re-fetch on error rows)",
        key="classify_url_col",
    )

refetch_errors = st.checkbox(
    "Re-fetch URLs for error/blank rows during classification",
    value=False,
    help="If the evidence column has an ERROR or is blank, try fetching the article again",
    key="refetch_errors",
)

classify_button = st.button("Classify Mentions", key="classify_btn", disabled=st.session_state.get("cl_running", False))

if classify_button:
    # ── Load source data ─────────────────────────────────────────────────────
    cl_headers: list[str] = []
    cl_rows: list[dict] = []

    if classify_source == "Use output from Step 1 above":
        if "enriched_rows" not in st.session_state:
            st.error("No Step 1 output found. Run Extract Mentions first or upload a file.")
            st.stop()
        cl_headers = list(st.session_state["enriched_headers"])
        cl_rows = [dict(r) for r in st.session_state["enriched_rows"]]
    else:
        if classify_file is None:
            st.error("Please upload a CSV or XLSX file.")
            st.stop()
        cl_headers, cl_rows = read_uploaded_table_filtered(
            classify_file, classify_selected_sheet,
            date_start=classify_date_start, date_end=classify_date_end,
        )
        if not cl_headers:
            st.error("Could not read the uploaded file. Make sure it's a valid CSV or XLSX with a header row.")
            st.stop()

    if not cl_rows:
        if classify_source != "Use output from Step 1 above" and classify_selected_sheet is not None:
            st.error(f"No rows found on sheet '{classify_selected_sheet}' after filtering.")
        else:
            st.error("File has no data rows.")
        st.stop()

    evidence_col = normalize_column(evidence_col_input, cl_headers)
    if evidence_col is None:
        st.error(f"Evidence column '{evidence_col_input}' not found. Available: {', '.join(cl_headers)}")
        st.stop()

    classify_url_col = normalize_column(classify_url_col_input, cl_headers)

    CLASS_COL = "Classification"
    JUST_COL = "Justification"
    CONF_COL = "Confidence"

    out_cl_headers = cl_headers[:]
    for col_name in (CLASS_COL, JUST_COL, CONF_COL):
        if col_name not in out_cl_headers:
            out_cl_headers.append(col_name)

    llm_config, llm_err = build_llm_config()
    if llm_err:
        st.error(llm_err)
        st.stop()

    eligible_indices = []
    for idx, row in enumerate(cl_rows):
        url = row.get(classify_url_col or "", "").strip() if classify_url_col else ""
        evidence = row.get(evidence_col, "").strip()
        if url.startswith("http") or (evidence and evidence != "NO_URL"):
            eligible_indices.append(idx)

    if not eligible_indices:
        st.error("No eligible rows found.")
        st.stop()

    # ── Snapshot into session_state ──────────────────────────────────────────
    st.session_state["cl_running"] = True
    st.session_state["cl_progress"] = 0
    st.session_state["cl_ok"] = 0
    st.session_state["cl_err"] = 0
    st.session_state["cl_rows"] = cl_rows
    st.session_state["cl_headers"] = cl_headers
    st.session_state["cl_out_headers"] = out_cl_headers
    st.session_state["cl_evidence_col"] = evidence_col
    st.session_state["cl_url_col"] = classify_url_col
    st.session_state["cl_eligible"] = eligible_indices
    st.session_state["cl_refetch"] = refetch_errors
    st.session_state["cl_llm_config"] = {k: v for k, v in llm_config.items() if k != "bedrock_client"}
    st.session_state["cl_results"] = {}  # step_index → {class, just, conf}
    st.rerun()

# ── Resume / run classification loop ─────────────────────────────────────────
if st.session_state.get("cl_running"):
    cl_rows = st.session_state["cl_rows"]
    out_cl_headers = st.session_state["cl_out_headers"]
    evidence_col = st.session_state["cl_evidence_col"]
    classify_url_col = st.session_state["cl_url_col"]
    eligible_indices = st.session_state["cl_eligible"]
    cl_refetch = st.session_state["cl_refetch"]
    cl_llm_config = st.session_state["cl_llm_config"]
    cl_saved_results = st.session_state["cl_results"]
    cl_start = st.session_state["cl_progress"]

    CLASS_COL = "Classification"
    JUST_COL = "Justification"
    CONF_COL = "Confidence"

    total_cl = len(eligible_indices)
    cl_ok = st.session_state["cl_ok"]
    cl_err = st.session_state["cl_err"]

    # Rebuild bedrock client if needed
    if cl_llm_config.get("provider") == "bedrock" and "bedrock_client" not in cl_llm_config:
        try:
            cl_llm_config["bedrock_client"] = boto3.client("bedrock-runtime", region_name=BEDROCK_REGION)
        except Exception as exc:
            st.error(f"Could not reconnect to AWS Bedrock: {exc}")
            st.session_state["cl_running"] = False
            st.stop()

    # Re-apply already-completed results to rows
    for done_step, saved in cl_saved_results.items():
        idx = eligible_indices[int(done_step)]
        cl_rows[idx][CLASS_COL] = saved["classification"]
        cl_rows[idx][JUST_COL] = saved["justification"]
        cl_rows[idx][CONF_COL] = saved["confidence"]

    if st.button("⛔ Cancel Classification", key="cancel_cl"):
        st.session_state["cl_running"] = False
        st.warning("Classification cancelled.")
        st.rerun()

    cl_status = st.empty()
    cl_progress = st.progress(cl_start / total_cl if total_cl else 0)
    cl_stats = st.empty()

    if cl_start > 0:
        cl_status.markdown(
            f"<small style='font-family:DM Mono,monospace;color:#888'>▶ Resuming from {cl_start+1}/{total_cl}…</small>",
            unsafe_allow_html=True,
        )

    for step in range(cl_start, total_cl):
        idx = eligible_indices[step]
        row = cl_rows[idx]
        evidence = row.get(evidence_col, "").strip()
        url = row.get(classify_url_col or "", "").strip() if classify_url_col else ""

        cl_status.markdown(
            f"<small style='font-family:DM Mono,monospace;color:#555'>🏷 Row {idx+1}: classifying…</small>",
            unsafe_allow_html=True,
        )

        is_error_or_blank = (
            not evidence
            or evidence == "NO_URL"
            or evidence.startswith("ERROR")
        )

        if is_error_or_blank and cl_refetch and url.startswith("http"):
            cl_status.markdown(
                f"<small style='font-family:DM Mono,monospace;color:#888'>🔄 Re-fetching {url[:60]}…</small>",
                unsafe_allow_html=True,
            )
            refetched = fetch_and_extract(url, row.get(HIT_SENTENCE_COL, ""))
            if refetched and not refetched.startswith("ERROR"):
                evidence = refetched
                is_error_or_blank = False

        if is_error_or_blank:
            row[CLASS_COL] = "E"
            if not evidence or evidence == "NO_URL":
                row[JUST_COL] = "No evidence text available and no URL to fetch."
            else:
                row[JUST_COL] = f"Evidence column returned an error: {evidence[:80]}"
            row[CONF_COL] = "90"
            cl_err += 1
        else:
            result = classify_mention(evidence, cl_llm_config, url=url)
            row[CLASS_COL] = result["classification"]
            row[JUST_COL] = result["justification"]
            row[CONF_COL] = str(result["confidence"])
            if result["classification"] == "E":
                cl_err += 1
            else:
                cl_ok += 1

            time.sleep(0.3)

        # ── Persist progress ─────────────────────────────────────────────────
        cl_saved_results[step] = {
            "classification": row[CLASS_COL],
            "justification": row[JUST_COL],
            "confidence": row[CONF_COL],
        }
        st.session_state["cl_progress"] = step + 1
        st.session_state["cl_ok"] = cl_ok
        st.session_state["cl_err"] = cl_err
        st.session_state["cl_results"] = cl_saved_results
        st.session_state["cl_rows"] = cl_rows

        cl_progress.progress((step + 1) / total_cl)
        cl_stats.markdown(
            f"""<div class="stat-row">
                <div class="stat-box"><div class="stat-num">{step+1}/{total_cl}</div><div class="stat-label">Classified</div></div>
                <div class="stat-box"><div class="stat-num">{cl_ok}</div><div class="stat-label">OK</div></div>
                <div class="stat-box"><div class="stat-num errors">{cl_err}</div><div class="stat-label">Errors</div></div>
            </div>""",
            unsafe_allow_html=True,
        )

    # ── Complete ─────────────────────────────────────────────────────────────
    cl_xlsx_bytes = rows_to_xlsx_bytes(out_cl_headers, cl_rows)

    cl_status.markdown(
        "<small style='font-family:DM Mono,monospace;color:#1a1a1a'>✓ Classification complete</small>",
        unsafe_allow_html=True,
    )

    st.markdown("<hr>", unsafe_allow_html=True)
    st.download_button(
        label="⬇ Download Classified XLSX",
        data=cl_xlsx_bytes,
        file_name="mentions_classified.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="download_classified",
    )

    st.session_state["cl_running"] = False