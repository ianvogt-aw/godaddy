"""
GoDaddy IC Data Grid Generator
================================
Pulls mentions from all 18 Cision One mention streams, tags T1 vs Non-T1,
and outputs an Excel workbook matching the IC Data grid format.

Usage:
    export CISION_API_TOKEN="your-token-here"
    python generate_ic_data_grid.py --after 2026-05-01 --before 2026-06-01

    # Append to an existing grid:
    python generate_ic_data_grid.py --after 2026-06-01 --before 2026-06-11 \
        --append-to GoDaddy_IC_Data_-_2026.xlsx

    # Also re-check Yellow-flagged mentions against the full article text
    # (fetches each flagged mention's own URL — off by default, adds runtime):
    python generate_ic_data_grid.py --after 2026-05-01 --before 2026-06-01 --fetch-full-text

    # Every newly-added row is LLM-reviewed by default (Bedrock Claude judges
    # whether it's a legit GoDaddy mention, using the full article text where
    # fetchable) — pass --no-llm-review to skip this and save time/cost:
    python generate_ic_data_grid.py --after 2026-05-01 --before 2026-06-01 --no-llm-review

Requirements:
    pip install requests openpyxl python-dateutil beautifulsoup4 lxml boto3

    Bedrock review (--llm-review, on by default) needs AWS credentials configured
    (e.g. via `aws configure` or standard AWS env vars) with Bedrock access:
        export BEDROCK_REGION="us-east-2"                                  # optional, this is the default
        export BEDROCK_MODEL_ID="us.anthropic.claude-sonnet-4-20250514-v1:0"  # optional, this is the default

    Optional, only used by --fetch-full-text as upgrades over plain `requests`
    (both are skipped silently if absent):
        export TAVILY_API_KEY="your-tavily-key"   # best at paywalls/bot-detection
        pip install curl_cffi                     # Chrome TLS-fingerprint spoofing
"""

import argparse
import os
import re
import sys
import time
import json
import requests
from copy import copy
from datetime import datetime, timedelta, timezone
from urllib.parse import urlparse
from collections import defaultdict, Counter

try:
    import boto3
except ImportError:
    boto3 = None

import openpyxl
from bs4 import BeautifulSoup
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter
from dateutil import parser as dtparser

# ============================================================================
# CONFIGURATION
# ============================================================================

BASE_URL = "https://api.cision.one"
REQUEST_INTERVAL = 7.0
MAX_RETRIES = 5

# Stream ID → metadata mapping
# Each entry: (stream_id, label, tier, category, tab_name, input_name_template)
STREAMS = [
    # ── PRODUCTS T1 ───────────────────────────────────────────────────
    {"id": 941746, "label": "AGI",      "tier": "T1",     "group": "PRODUCTS",  "tab": "AGI (Product) + Int.",       "input_name": "GoDaddy Product - T1 - AGI"},
    {"id": 941752, "label": "ANS",      "tier": "T1",     "group": "PRODUCTS",  "tab": "ANS (Product) + Int.",       "input_name": "GoDaddy Product - T1 - ANS"},
    {"id": 941764, "label": "Airo",     "tier": "T1",     "group": "PRODUCTS",  "tab": "Airo (Product) + Int.",      "input_name": "GoDaddy Product - T1 - Airo"},
    {"id": 941768, "label": "Commerce", "tier": "T1",     "group": "PRODUCTS",  "tab": "Commerce (Product) + Int.",  "input_name": "GoDaddy Product - T1 - Commerce"},
    {"id": 941776, "label": "Other",    "tier": "T1",     "group": "PRODUCTS",  "tab": "Other (Product) + Int.",     "input_name": "GoDaddy Product - T1 - Other"},
    # ── CORPORATE T1 ──────────────────────────────────────────────────
    {"id": 941782, "label": "Brand",               "tier": "T1",  "group": "CORPORATE", "tab": "Brand + Int.",               "input_name": "GoDaddy Brand - T1"},
    {"id": 941789, "label": "Finance",             "tier": "T1",  "group": "CORPORATE", "tab": "Finance + Int.",             "input_name": "GoDaddy - Finance - T1"},
    {"id": 941792, "label": "Thought Leadership",  "tier": "T1",  "group": "CORPORATE", "tab": "Thought Leadership + Int.",  "input_name": "GoDaddy Thought Leadership - T1", "split_by_person": True},
    {"id": 964987, "label": "ANS Open Standard",   "tier": "T1",  "group": "ANS_OPEN_STANDARD", "tab": "ANS Open Standard", "input_name": "ANS Open Standard - T1"},
    {"id": 965728, "label": "Brand Identity",      "tier": "T1",  "group": "BRAND_IDENTITY", "tab": "Brand Identity", "input_name": "Brand Identity - T1"},
    # ── SBRL T1 ───────────────────────────────────────────────────────
    {"id": 941796, "label": "sbrl", "tier": "T1",     "group": "SBRL", "tab": "GDSBRL + Int.", "input_name": "GoDaddy Venture Forward - T1"},
    # ── PRODUCTS (Non-T1) ─────────────────────────────────────────────
    {"id": 943540, "label": "AGI",      "tier": "Non-T1", "group": "PRODUCTS",  "tab": "AGI (Product) + Int.",       "input_name": "GoDaddy Product - Non T1 - AGI"},
    {"id": 943645, "label": "ANS",      "tier": "Non-T1", "group": "PRODUCTS",  "tab": "ANS (Product) + Int.",       "input_name": "GoDaddy Product - Non T1 - ANS"},
    {"id": 943635, "label": "Airo",     "tier": "Non-T1", "group": "PRODUCTS",  "tab": "Airo (Product) + Int.",      "input_name": "GoDaddy Product - Non T1 - Airo"},
    {"id": 943636, "label": "Commerce", "tier": "Non-T1", "group": "PRODUCTS",  "tab": "Commerce (Product) + Int.",  "input_name": "GoDaddy Product - Non T1- Commerce"},
    {"id": 943638, "label": "Other",    "tier": "Non-T1", "group": "PRODUCTS",  "tab": "Other (Product) + Int.",     "input_name": "GoDaddy Product - Non T1 - Other"},
    # ── CORPORATE (Non-T1) ────────────────────────────────────────────
    {"id": 943639, "label": "Brand",               "tier": "Non-T1", "group": "CORPORATE", "tab": "Brand + Int.",               "input_name": "GoDaddy Brand - Non T1"},
    {"id": 943641, "label": "Finance",             "tier": "Non-T1", "group": "CORPORATE", "tab": "Finance + Int.",             "input_name": "GoDaddy - Finance - Non T1"},
    {"id": 943642, "label": "Thought Leadership",  "tier": "Non-T1", "group": "CORPORATE", "tab": "Thought Leadership + Int.",  "input_name": "GoDaddy Thought Leadership - Non T1", "split_by_person": True},
    {"id": 964967, "label": "ANS Open Standard",   "tier": "Non-T1", "group": "ANS_OPEN_STANDARD", "tab": "ANS Open Standard", "input_name": "ANS Open Standard - Non T1"},
    {"id": 965727, "label": "Brand Identity",      "tier": "Non-T1", "group": "BRAND_IDENTITY", "tab": "Brand Identity", "input_name": "Brand Identity - Non T1"},
    # ── SBRL (Non-T1) ────────────────────────────────────────────────
    {"id": 943644, "label": "sblr", "tier": "Non-T1", "group": "SBRL", "tab": "GDSBRL + Int.", "input_name": "GoDaddy Venture Forward - Non T1"},
]

# Thought Leadership mentions are split out of the combined stream into one tab per
# spokesperson, based on which person's name appears in the mention's Keywords field.
# A mention tagged to multiple people (e.g. "Travis Muhlestein;Jared Sine") lands in
# both tabs.
PEOPLE = [
    "Aman Bhutani",
    "Gourav Pani",
    "Mark McCaffrey",
    "Kasturi Mudulodu",
    "Jared Sine",
    "Sarfraz Nakai",
    "Charles Beadnall",
    "Travis Muhlestein",
    "Dimitria Elmore",
    "Paul Bindel",
    "Phontip Palitwanon",
]
PEOPLE_TABS = {name: f"{name} + Int." for name in PEOPLE}
PEOPLE_LOOKUP = {name.lower(): tab for name, tab in PEOPLE_TABS.items()}

# ANS Open Standard mentions aren't fundamentally about "is this a legit GoDaddy
# mention" the same way the rest of the grid is — skip both the GoDaddy-focused
# Hit Sentence rewrite and the LLM legitimacy review for this tab.
NO_GODADDY_FOCUS_TABS = {"ANS Open Standard"}

# Desired tab ordering in the output workbook
TAB_ORDER = [
    "GDSBRL + Int.",
    "Commerce (Product) + Int.",
    "AGI (Product) + Int.",
    "Airo (Product) + Int.",
    "ANS (Product) + Int.",
    "Other (Product) + Int.",
    "ANS Open Standard",
    "Brand + Int.",
    "Finance + Int.",
] + list(PEOPLE_TABS.values()) + [
    "Brand Identity",  # left uncolored (no entry in TAB_COLOR_BY_TAB) — sits at the end
]

# The 42 columns matching the existing IC Data grid, plus 4 LLM-review columns
# (AQ-AT) this script owns — populated for newly-added rows when the Bedrock
# review step (--llm-review, on by default) runs, blank otherwise.
IC_COLUMNS = [
    "Date", "Time", "Document ID", "URL", "Input Name", "Keywords",
    "Information Type", "Source Type", "Source Name", "Source Domain",
    "Content Type", "Author Name", "Author Handle", "Title",
    "Opening Text", "Hit Sentence", "Image", "Hashtags", "Links",
    "Country", "Region", "State", "City", "Language", "Sentiment",
    "Keyphrases", "Reach", "AVE", "Social Echo", "Editorial Echo",
    "Engagement", "Shares", "Quotes", "Likes", "Replies", "Reposts",
    "Comments", "Reactions", "Views", "Estimated Views",
    "Document Tags", "Custom Categories",
    "Review Evidence", "Review Classification", "Review Justification", "Review Confidence",
]

# Column widths for readability
COL_WIDTHS = {
    "A": 12, "B": 8, "C": 16, "D": 50, "E": 35, "F": 30,
    "G": 14, "H": 14, "I": 25, "J": 25, "K": 15, "L": 18,
    "M": 15, "N": 60, "O": 40, "P": 60, "Q": 40, "R": 20,
    "S": 40, "T": 15, "U": 15, "V": 15, "W": 15, "X": 12,
    "Y": 10, "Z": 30, "AA": 12, "AB": 12, "AC": 12, "AD": 12,
    "AE": 12, "AF": 10, "AG": 10, "AH": 10, "AI": 10, "AJ": 10,
    "AK": 10, "AL": 10, "AM": 10, "AN": 12, "AO": 15, "AP": 25,
    "AQ": 60, "AR": 14, "AS": 50, "AT": 14,
}

# Fixed column indices (0-based) used both when building a row and when reading
# one back — kept as named lookups instead of magic numbers/[-1] so adding
# trailing columns (like the review ones above) never silently shifts meaning.
URL_COL_IDX = IC_COLUMNS.index("URL")
HIT_SENTENCE_COL_IDX = IC_COLUMNS.index("Hit Sentence")
REVIEW_EVIDENCE_COL_IDX = IC_COLUMNS.index("Review Evidence")
REVIEW_CLASSIFICATION_COL_IDX = IC_COLUMNS.index("Review Classification")
REVIEW_JUSTIFICATION_COL_IDX = IC_COLUMNS.index("Review Justification")
REVIEW_CONFIDENCE_COL_IDX = IC_COLUMNS.index("Review Confidence")


# ============================================================================
# API CLIENT (rate-limited)
# ============================================================================

class CisionOneClient:
    def __init__(self, api_token: str):
        self.token = api_token
        self.session = requests.Session()
        self.session.headers.update({
            "X-Auth-Token": self.token,
            "Accept": "application/json",
        })
        self._last_request_time = 0.0

    def _throttled_get(self, url, params):
        for attempt in range(1, MAX_RETRIES + 1):
            elapsed = time.time() - self._last_request_time
            if elapsed < REQUEST_INTERVAL:
                time.sleep(REQUEST_INTERVAL - elapsed)
            self._last_request_time = time.time()
            resp = self.session.get(url, params=params)
            if resp.status_code != 429:
                return resp
            retry_after = resp.headers.get("Retry-After")
            wait = float(retry_after) if retry_after else REQUEST_INTERVAL * (2 ** (attempt - 1))
            print(f"    ⏳ Rate-limited (429). Retry {attempt}/{MAX_RETRIES} — waiting {wait:.0f}s")
            time.sleep(wait)
        resp.raise_for_status()

    def get_mentions(self, stream_id, after, before, page=1, page_size=500):
        resp = self._throttled_get(
            f"{BASE_URL}/public/api/v2/mentions/{stream_id}",
            {
                "filter[range][after]": after,
                "filter[range][before]": before,
                "pagination[page]": page,
                "pagination[page_size]": page_size,
                "sort[field]": "timestamp",
                "sort[order]": "desc",
                "format": "json",
            },
        )
        resp.raise_for_status()
        return resp.json()

    def get_all_mentions(self, stream_id, after, before, page_size=500, max_pages=9):
        all_mentions = []
        page = 1
        while page <= max_pages and (page * page_size) <= 5000:
            try:
                batch = self.get_mentions(stream_id, after, before, page=page, page_size=page_size)
            except requests.HTTPError as e:
                if e.response.status_code == 400:
                    print(f"    ⚠️  Hit API ceiling at page {page} — returning {len(all_mentions)} mentions")
                    break
                raise
            if not batch:
                break
            all_mentions.extend(batch)
            if len(batch) < page_size:
                break
            page += 1
        return all_mentions

    def get_all_mentions_chunked(self, stream_id, after, before, chunk_days=1, page_size=500):
        start = dtparser.isoparse(after)
        end = dtparser.isoparse(before)
        chunk = timedelta(days=chunk_days)
        all_mentions = []
        window_start = start
        while window_start < end:
            window_end = min(window_start + chunk, end)
            w_after = window_start.strftime("%Y-%m-%dT%H:%M:%S.000Z")
            w_before = window_end.strftime("%Y-%m-%dT%H:%M:%S.000Z")
            batch = self.get_all_mentions(stream_id, w_after, w_before, page_size=page_size)
            all_mentions.extend(batch)
            window_start = window_end
        # Deduplicate
        seen = set()
        deduped = []
        for m in all_mentions:
            mid = m.get("id")
            if mid not in seen:
                seen.add(mid)
                deduped.append(m)
        return deduped


# ============================================================================
# FIELD MAPPING: Cision One API → IC Data Grid columns
# ============================================================================

def extract_domain(url: str) -> str:
    """Extract domain from URL."""
    if not url:
        return ""
    try:
        parsed = urlparse(url)
        domain = parsed.netloc or parsed.path.split("/")[0]
        return domain.replace("www.", "")
    except Exception:
        return ""


def map_medium_to_source_type(medium: str, mention_type: str) -> str:
    """Map API medium/type to IC Data 'Source Type' values."""
    mapping = {
        "Online": "online news",
        "Print": "print",
        "TV": "tv",
        "Radio": "radio",
        "Social": "social network",
        "Podcast": "podcast",
        "Magazine": "magazine",
    }
    return mapping.get(medium, mention_type or "")


def map_type_to_content_type(mention_type: str) -> str:
    """Map API 'type' to IC Data 'Content Type' values."""
    mapping = {
        "onlineArticle": "News Article",
        "printArticle": "News Article",
        "tvClip": "Video",
        "radioClip": "Audio",
        "socialPost": "Social Post",
        "podcastEpisode": "Audio",
        "magazineArticle": "News Article",
        "blogPost": "Blog Post",
    }
    return mapping.get(mention_type, mention_type or "")


def map_info_type(medium: str) -> str:
    """Map API medium to IC Data 'Information Type'."""
    if medium in ("TV", "Radio"):
        return "broadcast"
    elif medium == "Social":
        return "social"
    return "news"


def map_sentiment_label(score) -> str:
    """Map numeric sentiment score to label."""
    if score is None:
        return ""
    if score > 0.1:
        return "positive"
    elif score < -0.1:
        return "negative"
    return "neutral"


def map_language_code(code: str) -> str:
    """Map ISO language codes to display names."""
    if not code:
        return ""
    lang_map = {
        "en": "English", "en-US": "English", "en-GB": "English", "en-AU": "English",
        "es": "Spanish", "es-ES": "Spanish", "es-MX": "Spanish",
        "fr": "French", "fr-FR": "French", "fr-CA": "French",
        "de": "German", "de-DE": "German",
        "pt": "Portuguese", "pt-BR": "Portuguese",
        "it": "Italian", "ja": "Japanese", "ko": "Korean",
        "zh": "Chinese", "ar": "Arabic", "hi": "Hindi",
        "nl": "Dutch", "sv": "Swedish", "da": "Danish",
        "no": "Norwegian", "fi": "Finnish", "pl": "Polish",
        "ru": "Russian", "tr": "Turkish", "th": "Thai",
        "id": "Indonesian", "ms": "Malay", "vi": "Vietnamese",
    }
    return lang_map.get(code, lang_map.get(code.split("-")[0], code))


_SENTENCE_SPLIT_RE = re.compile(r"(?<=[.!?])\s+")


def focus_hit_sentence(text: str, keyword: str = "GoDaddy") -> tuple[str, bool]:
    """Prefer the sentence mentioning `keyword` over Cision's raw excerpt/transcript
    blob. The API only gives us one fixed excerpt with no keyword offsets or full
    article body, so this can only re-center within text Cision already returned —
    if `keyword` isn't present anywhere in it, the text is returned unchanged.
    Returns (text, found) so callers can flag mentions where `keyword` never
    turned up at all."""
    if not text:
        return text, False
    found = keyword.lower() in text.lower()
    if not found:
        return text, False
    for sentence in _SENTENCE_SPLIT_RE.split(text):
        if keyword.lower() in sentence.lower():
            return sentence.strip(), True
    return text, True


# ============================================================================
# FULL-ARTICLE FALLBACK (--fetch-full-text)
# ============================================================================
# Cision's excerpt is a single fixed snippet — if "GoDaddy" isn't in it, that
# doesn't mean the article never mentions GoDaddy, just that Cision's snippet
# didn't happen to land there. For mentions flagged Yellow, we can fetch the
# article's own public URL and search the real full text ourselves, the same
# way mention_review_streamlit/app.py does for old Meltwater data (fetch ->
# strip to body text -> split into sentences -> find the one with "GoDaddy").

ARTICLE_FETCH_TIMEOUT = 15
ARTICLE_FETCH_USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36"
)
ARTICLE_FETCH_DELAY = 1.0  # seconds between fetches, so we don't hammer news sites

# Optional, in priority order: Tavily's Extract API (best at paywalls/bot-detection,
# no local browser needed — set TAVILY_API_KEY to enable) and curl_cffi (mimics
# Chrome's TLS fingerprint, gets past a lot of basic bot-blocking WAFs). Both are
# skipped silently if unconfigured/not installed — plain `requests` always runs last.
TAVILY_API_KEY = os.environ.get("TAVILY_API_KEY", "")
TAVILY_EXTRACT_URL = "https://api.tavily.com/extract"

_ABBREV_RE = re.compile(r"\b(Mr|Mrs|Ms|Dr|Prof|Sr|Jr|vs|etc|approx|Inc|Ltd|Corp|Co)\.")
_ROBUST_SENTENCE_SPLIT_RE = re.compile(r"(?<=[.!?])\s+(?=[A-Z\"'‘’“”\(\[])")


def _split_sentences_robust(text: str) -> list[str]:
    """Sentence splitter for full article bodies — tolerant of abbreviations
    ("Inc.", "Corp.", etc.) that would otherwise cause false splits. Ported from
    mention_review_streamlit/app.py's split_sentences()."""
    text = re.sub(r"\s+", " ", text).strip()
    text = _ABBREV_RE.sub(r"\1<DOT>", text)
    parts = _ROBUST_SENTENCE_SPLIT_RE.split(text)
    return [p.replace("<DOT>", ".").strip() for p in parts if p.strip()]


def _extract_article_text(html: str) -> str:
    """Best-effort article body extraction. Ported from
    mention_review_streamlit/app.py's extract_article_text()."""
    soup = BeautifulSoup(html, "lxml")
    for tag in soup(["script", "style", "nav", "header", "footer", "aside",
                     "noscript", "iframe", "form"]):
        tag.decompose()
    for selector in [
        "article", "main", '[role="main"]', ".post-content", ".entry-content",
        ".article-body", ".story-body", ".article__body", ".content-body",
        "#article-body", ".post-body",
    ]:
        container = soup.select_one(selector)
        if container:
            text = container.get_text(separator=" ", strip=True)
            if len(text) > 200:
                return text
    body = soup.find("body")
    return (body or soup).get_text(separator=" ", strip=True)


def _fetch_via_tavily(url: str) -> str | None:
    """Fetch article text via Tavily's Extract API. Returns None if TAVILY_API_KEY
    isn't set or the call fails — this is a silent, optional upgrade."""
    if not TAVILY_API_KEY:
        return None
    try:
        resp = requests.post(
            TAVILY_EXTRACT_URL,
            json={"urls": [url], "extract_depth": "advanced", "format": "text"},
            headers={"Authorization": f"Bearer {TAVILY_API_KEY}", "Content-Type": "application/json"},
            timeout=30,
        )
        resp.raise_for_status()
        results = resp.json().get("results", [])
        if results:
            return results[0].get("raw_content", "") or results[0].get("content", "") or None
        return None
    except Exception:
        return None


def _fetch_via_curl_cffi(url: str) -> str | None:
    """Fetch via curl_cffi (Chrome TLS-fingerprint impersonation) if it's installed.
    Returns None if the package is missing or the fetch fails."""
    try:
        from curl_cffi import requests as cffi_requests
    except ImportError:
        return None
    try:
        resp = cffi_requests.get(url, impersonate="chrome120", timeout=ARTICLE_FETCH_TIMEOUT, allow_redirects=True)
        if resp.status_code != 200:
            return None
        return _extract_article_text(resp.text)
    except Exception:
        return None


def fetch_full_article_text(url: str) -> str | None:
    """Best-effort fetch of the full article body at `url`, trying (in order)
    Tavily's Extract API, curl_cffi, then plain `requests`. Returns None if every
    method fails — callers must fall back gracefully; this is a bonus check, not
    something the pipeline depends on."""
    text = _fetch_via_tavily(url)
    if text:
        return text

    text = _fetch_via_curl_cffi(url)
    if text:
        return text

    try:
        resp = requests.get(
            url,
            headers={"User-Agent": ARTICLE_FETCH_USER_AGENT, "Accept-Language": "en-US,en;q=0.9"},
            timeout=ARTICLE_FETCH_TIMEOUT,
            allow_redirects=True,
        )
        if resp.status_code != 200:
            return None
        return _extract_article_text(resp.text)
    except Exception:
        return None


def find_keyword_sentence(text: str, keyword: str = "GoDaddy") -> str | None:
    """Locate the first sentence in a full article body mentioning `keyword`.
    Returns None if it isn't there either."""
    if not text or keyword.lower() not in text.lower():
        return None
    for sentence in _split_sentences_robust(text):
        if keyword.lower() in sentence.lower():
            return sentence.strip()
    return None


class Row(list):
    """A data row that also carries metadata alongside its 46 IC_COLUMNS values:
    is_t1 is the single source of truth for both the Custom Categories value and
    the base Orange/Pink fill (see mention_to_row/_row_fill). hit_sentence_flagged
    records whether focus_hit_sentence had to give up on finding "GoDaddy" in
    Cision's excerpt — it only affects Hit Sentence content, not row color."""
    is_t1 = False
    hit_sentence_flagged = False


def mention_to_row(mention: dict, stream_config: dict, fetch_full_text: bool = False) -> list:
    """Convert a single API mention + stream metadata → 46-element list matching
    IC_COLUMNS. If fetch_full_text is set and Cision's excerpt doesn't mention
    "GoDaddy", makes one best-effort fetch of the mention's own public URL and
    searches the real article body before giving up on the Hit Sentence rewrite."""
    is_t1 = stream_config["tier"] == "T1"

    pub = mention.get("publishedAt", "")
    dt = None
    if pub:
        try:
            dt = dtparser.isoparse(pub)
        except Exception:
            pass

    medium = mention.get("medium", "")
    mtype = mention.get("type", "")
    social = mention.get("social") or {}
    social_echo = sum(v for v in [
        social.get("x"), social.get("facebook"),
        social.get("reddit"), social.get("pinterest"),
    ] if v is not None) or None

    keywords_raw = mention.get("keywords") or []
    keywords_str = ";".join(keywords_raw) if keywords_raw else ""

    content_text = mention.get("excerpt") or mention.get("transcript") or ""
    hit_sentence_flagged = False
    if stream_config["tab"] not in NO_GODADDY_FOCUS_TABS:
        content_text, keyword_found = focus_hit_sentence(content_text, "GoDaddy")
        if not keyword_found and fetch_full_text and mention.get("url"):
            time.sleep(ARTICLE_FETCH_DELAY)
            full_text = fetch_full_article_text(mention["url"])
            better_sentence = find_keyword_sentence(full_text, "GoDaddy") if full_text else None
            if better_sentence:
                content_text = better_sentence
                keyword_found = True
        hit_sentence_flagged = not keyword_found

    # Fall back to Cision's internal link when the mention has no public URL
    # (common for broadcast clips and some social posts).
    url = mention.get("url") or mention.get("internalLink") or ""

    row = Row([
        dt.date() if dt else None,                         # Date
        dt.time() if dt else None,                         # Time
        mention.get("id"),                                  # Document ID
        url,                                                 # URL
        stream_config["input_name"],                        # Input Name
        keywords_str,                                       # Keywords
        map_info_type(medium),                              # Information Type
        map_medium_to_source_type(medium, mtype),           # Source Type
        mention.get("source", ""),                          # Source Name
        extract_domain(url),                                # Source Domain
        map_type_to_content_type(mtype),                    # Content Type
        mention.get("author", ""),                          # Author Name
        None,                                               # Author Handle (not in API)
        mention.get("title", ""),                           # Title
        None,                                               # Opening Text (not reliably in API)
        content_text,                                       # Hit Sentence
        None,                                               # Image (not in API)
        None,                                               # Hashtags (not in API)
        None,                                               # Links (not in API)
        mention.get("locationCountry", ""),                 # Country
        None,                                               # Region (not in API)
        mention.get("locationState", ""),                   # State
        mention.get("locationCity", ""),                     # City
        map_language_code(mention.get("languageCode", "")), # Language
        map_sentiment_label(mention.get("sentiment")),      # Sentiment
        keywords_str,                                       # Keyphrases (same as Keywords)
        mention.get("audience"),                            # Reach
        mention.get("advertisingValue"),                    # AVE
        social_echo,                                        # Social Echo
        None,                                               # Editorial Echo (not in API)
        None,                                               # Engagement (not in API)
        None,                                               # Shares
        None,                                               # Quotes
        None,                                               # Likes
        None,                                               # Replies
        None,                                               # Reposts
        None,                                               # Comments
        None,                                               # Reactions
        None,                                               # Views
        None,                                               # Estimated Views
        None,                                               # Document Tags (not in API)
        "Y" if is_t1 else None,                              # Custom Categories: Y for orange (T1) rows, blank for pink
        None,                                               # Review Evidence (filled in by review_new_rows)
        None,                                               # Review Classification
        None,                                               # Review Justification
        None,                                               # Review Confidence
    ])
    row.hit_sentence_flagged = hit_sentence_flagged
    row.is_t1 = is_t1
    return row


# ============================================================================
# LLM MENTION REVIEW (--llm-review, on by default)
# ============================================================================
# Reproduces the mention_review_streamlit Classify Mentions step, but folded
# directly into this pipeline and run automatically on every newly-added row —
# no separate app, no manual CSV round-trip. Unlike --fetch-full-text (which
# only kicks in for Yellow-flagged rows and just fixes the Hit Sentence text),
# this always tries to fetch the mention's full article body and hands as much
# of it as possible to Claude to judge whether it's a legit GoDaddy mention.

BEDROCK_REGION = os.environ.get("BEDROCK_REGION", "us-east-2")
BEDROCK_MODEL_ID = os.environ.get("BEDROCK_MODEL_ID", "us.anthropic.claude-sonnet-4-20250514-v1:0")

# How much evidence text we'll both send to Claude and store in the Review
# Evidence cell. Generous enough to cover nearly any news article in full.
EVIDENCE_CHAR_CAP = 12000

PRESS_RELEASE_DOMAINS = {
    "prnewswire.com", "businesswire.com", "globenewswire.com", "prweb.com",
    "newswire.com", "einpresswire.com", "accesswire.com", "prlog.org",
    "send2press.com", "24-7pressrelease.com", "newswire.ca", "marketwired.com",
    "marketwire.com", "presswire.com", "prfire.co.uk", "cisionone.com", "cision.com",
}

# Adapted from mention_review_streamlit/app.py's CLASSIFICATION_SYSTEM_PROMPT —
# same taxonomy and decision rules, just describing full-article evidence
# instead of a Before/Mention/After snippet.
REVIEW_SYSTEM_PROMPT = """\
You are classifying a news article's mention of "GoDaddy" based on the evidence text provided.

The evidence is either the full body text of the article (when we could fetch it) or a shorter \
fallback snippet from Cision (when we couldn't). "GoDaddy" may appear anywhere in the evidence —
read the whole thing, not just the beginning.

Classify into exactly ONE category:

A - GoDaddy Products: The evidence explicitly mentions a specific GoDaddy product, service, \
platform, tool, plan, feature, pricing tier, or product use case. \
Examples: "ANS", "Agent Name Service", "Airo", \
"registered through GoDaddy", "hosted on GoDaddy", "GoDaddy Economy", \
"GoDaddy Website Builder", "GoDaddy-issued certificate", domain sold at auction by GoDaddy, \
GoDaddy as registrar, host, DNS, WHOIS, SSL, email, payments, or security provider, \
"GoDaddy website team", "GoDaddy Secret Manager".

B - GoDaddy Brand/Company: GoDaddy is mentioned only at the company or brand level with no \
concrete offering or service use case, and the evidence is not about research/data. \
Examples: employee/executive mention, stock/investor/earnings coverage, GoDaddy named as \
competitor, sponsorships, partnerships, general company references, GoDaddy quote not \
tied to research output.

C - GoDaddy Research/Data: The evidence explicitly refers to GoDaddy Small Business Research Lab, \
GoDaddy research, GoDaddy data, GoDaddy survey findings, GoDaddy reports, GoDaddy indices, \
Venture Forward, or other research/data output from the Small Business Research Lab.

D - No GoDaddy mention: The evidence contains no explicit mention of "GoDaddy" at all.

E - Error: The evidence could not be read (blank, error, inaccessible).

F - Unclear: The evidence is readable but too thin, mixed, incomplete, or contradictory \
to classify confidently.

G - Press Release: The source article is a press release or wire-service distribution. \
Key signals: content begins with or contains "FOR IMMEDIATE RELEASE"; a press-release \
dateline format (e.g. "CITY, State, Month Day, Year —"); boilerplate sections like \
"About [Company]", "Media Contact:", or a "###" end marker; highly promotional \
third-person copy announcing a product launch, partnership, or executive quote in a \
way typical of corporate press releases. Classify as G regardless of whether GoDaddy \
is also a product/brand/research mention — press releases are discarded entirely.

H - Commercial/Advertisement: The mention is about or directly references a GoDaddy \
TV commercial, advertisement, ad campaign, promotional spot, or paid/sponsored content. \
Key signals: article reviews, recaps, or discusses a GoDaddy ad or commercial \
(e.g. Super Bowl spot, streaming ad); evidence references GoDaddy advertising  \
talent or a spokesperson in an advertising context (e.g. "Walton Goggins", \
"Goggle Glasses", a celebrity appearing in a GoDaddy ad); coverage of a GoDaddy \
ad campaign, ad creative, or marketing stunt rather than the company or its products. \
Classify as H regardless of other signals — commercial mentions are discarded.

Strict decision rules:
- Check for press release signals FIRST. If the source is clearly a press release, \
classify as G regardless of A/B/C signals.
- Check for commercial/ad signals SECOND. If the mention is about a GoDaddy commercial \
or ad campaign, classify as H regardless of A/B/C signals.
- If both A and B signals appear, choose A.
- If both B and C signals appear and the mention is about research/survey/report/index/data, choose C.
- Do NOT use A just because the article topic is domains, hosting, or websites. \
GoDaddy itself must be the provider, tool, service, or platform in the evidence sentence.
- Do NOT use B when the evidence is about GoDaddy research, survey, report, index, or data. Use C.
- Do NOT use D if "GoDaddy" appears anywhere in the evidence text.
- If the article is mainly about another company or topic, but the evidence says the \
domain/hosting/certificate/platform is from GoDaddy, that is still A.
- For stock, investor, earnings, or company-profile stories: if no specific GoDaddy offering \
is named in the evidence, use B.
- For non-English text, "GoDaddy" usually still appears in Latin letters.

Respond with ONLY valid JSON. No markdown fences, no explanation, no preamble:
{"classification": "X", "justification": "one short sentence naming the specific evidence", "confidence": NN}

Justification rules:
- Write exactly one short sentence.
- Name the specific evidence that led to the classification.
- Good: "Evidence says the domain was registered through GoDaddy."
- Good: "Evidence cites GoDaddy Small Business Research Lab data on microbusiness trends."
- Good: "Evidence lists GoDaddy stock as an S&P 500 mover but no product is discussed."
- Good: "Content contains 'FOR IMMEDIATE RELEASE' and a press-release dateline, indicating a wire-service distribution."
- Bad: "GoDaddy is mentioned." (too vague)

Confidence scoring:
- 95-100: evidence is direct and explicit
- 85-94: evidence is clear but slightly indirect or brief
- 70-84: evidence supports the label but some ambiguity exists
- 50-69: evidence is limited or mixed
- Below 50: substantial uncertainty
"""


def build_bedrock_client():
    """Build a Bedrock Runtime client for --llm-review. Returns None (with a
    printed warning) if boto3 isn't installed or AWS credentials/access aren't
    set up — callers should treat that as "review unavailable this run", not
    a fatal error."""
    if boto3 is None:
        print("⚠️  boto3 is not installed — skipping LLM review (pip install boto3, or pass --no-llm-review)")
        return None
    try:
        client = boto3.client("bedrock-runtime", region_name=BEDROCK_REGION)
        return client
    except Exception as e:
        print(f"⚠️  Could not create a Bedrock client — skipping LLM review: {e}")
        return None


def call_bedrock_claude(user_message: str, system_prompt: str, bedrock_client, model_id: str, max_tokens: int = 256) -> str:
    body = {
        "anthropic_version": "bedrock-2023-05-31",
        "max_tokens": max_tokens,
        "messages": [{"role": "user", "content": user_message}],
    }
    if system_prompt:
        body["system"] = system_prompt
    resp = bedrock_client.invoke_model(
        modelId=model_id,
        contentType="application/json",
        accept="application/json",
        body=json.dumps(body),
    )
    return json.loads(resp["body"].read())["content"][0]["text"].strip()


def classify_mention_full_text(evidence: str, bedrock_client, model_id: str, url: str = "") -> dict:
    """Send `evidence` (as much of the full article as we could get) to Claude
    and return {"classification": "X", "justification": "...", "confidence": N}."""
    if url:
        netloc = urlparse(url).netloc.lower().lstrip("www.")
        if any(netloc == d or netloc.endswith("." + d) for d in PRESS_RELEASE_DOMAINS):
            return {
                "classification": "G",
                "justification": f"Article URL is from a known press wire service ({netloc}).",
                "confidence": 98,
            }

    if not evidence.strip():
        return {"classification": "E", "justification": "No evidence text available.", "confidence": 90}

    try:
        raw = call_bedrock_claude(
            f"Classify this evidence:\n\n{evidence}",
            REVIEW_SYSTEM_PROMPT,
            bedrock_client,
            model_id,
        )
        raw = re.sub(r"^```(?:json)?\s*", "", raw)
        raw = re.sub(r"\s*```$", "", raw)
        result = json.loads(raw)
        if result.get("classification") not in ("A", "B", "C", "D", "E", "F", "G", "H"):
            result["classification"] = "F"
        result["confidence"] = int(result.get("confidence", 50))
        return result
    except Exception as e:
        return {"classification": "E", "justification": f"LLM classification call failed: {e}", "confidence": 90}


def review_new_rows(rows: list, bedrock_client, model_id: str, tab_name: str = "") -> None:
    """LLM-review each newly-added row for whether it's a legit GoDaddy mention.
    Fetches the mention's own public URL for the full article text where
    possible (falling back to Hit Sentence), then mutates each Row in place —
    writing Review Evidence/Classification/Justification/Confidence into the
    slots mention_to_row already reserved for them."""
    if not rows or bedrock_client is None:
        return

    print(f"  🏷  Reviewing {len(rows)} new mention(s) in '{tab_name}'…")
    counts = Counter()
    for row in rows:
        url = (row[URL_COL_IDX] or "").strip()
        full_text = fetch_full_article_text(url) if url.startswith("http") else None
        evidence = (full_text or row[HIT_SENTENCE_COL_IDX] or "").strip()[:EVIDENCE_CHAR_CAP]

        result = classify_mention_full_text(evidence, bedrock_client, model_id, url=url)
        row[REVIEW_EVIDENCE_COL_IDX] = evidence
        row[REVIEW_CLASSIFICATION_COL_IDX] = result["classification"]
        row[REVIEW_JUSTIFICATION_COL_IDX] = result["justification"]
        row[REVIEW_CONFIDENCE_COL_IDX] = result["confidence"]
        counts[result["classification"]] += 1
        time.sleep(0.3)

    summary = " ".join(f"{k}:{v}" for k, v in sorted(counts.items()))
    print(f"     {summary}")


# Classification -> destination tab. Rows landing on any other letter (D/E/F/G/H)
# stay put — D/G/H additionally get recolored Yellow by _row_fill.
RELOCATE_TARGETS = {
    "A": "AGI (Product) + Int.",
    "B": "Brand + Int.",
    "C": "GDSBRL + Int.",
}

# A Class A result only moves to AGI if it isn't already sitting in *some*
# Product tab — a mention already correctly bucketed under ANS/Airo/Commerce/
# Other shouldn't get yanked into AGI just because that's the generic Class A
# destination. Same idea for Class B and the Corporate tabs.
PRODUCT_TABS = {
    "AGI (Product) + Int.",
    "ANS (Product) + Int.",
    "Airo (Product) + Int.",
    "Commerce (Product) + Int.",
    "Other (Product) + Int.",
}
CORPORATE_TABS = {"Brand + Int.", "Finance + Int."} | set(PEOPLE_TABS.values())

# Tabs the LLM mention-review step never touches at all — no fetch, no classify,
# no relocation. Distinct from NO_GODADDY_FOCUS_TABS (which only gates the
# Hit Sentence rewrite in mention_to_row).
NO_LLM_REVIEW_TABS = {"Finance + Int.", "ANS Open Standard", "Brand Identity"} | set(PEOPLE_TABS.values())


def _relocation_target(tab_name: str, classification: str) -> str | None:
    """Where a classified row should move, or None to leave it in tab_name."""
    if classification == "A" and tab_name in PRODUCT_TABS:
        return None
    if classification == "B" and tab_name in CORPORATE_TABS:
        return None
    return RELOCATE_TARGETS.get(classification)


def review_and_relocate(tab_data: dict[str, list], bedrock_client, model_id: str) -> None:
    """Mutates tab_data in place. Per requirement: Orange (T1) rows are never
    reviewed or moved — they're left exactly where Cision put them for a human to
    check manually. Pink (Non-T1) rows are LLM-reviewed; classifications A/B/C
    relocate the row into the AGI/Brand/GDSBRL tab respectively, unless it's
    already in an equivalent Product/Corporate tab (see _relocation_target).
    D/E/F/G/H rows stay in their original tab. Skips NO_LLM_REVIEW_TABS entirely."""
    if bedrock_client is None:
        return

    for tab_name in list(tab_data.keys()):
        if tab_name in NO_LLM_REVIEW_TABS:
            continue
        rows = tab_data[tab_name]
        # Exclude Orange rows, and any row already classified in an earlier
        # iteration of this same loop (i.e. it was just relocated *into* this
        # tab) — otherwise a moved row would get reviewed a second time.
        pink_rows = [
            r for r in rows
            if not getattr(r, "is_t1", False) and r[REVIEW_CLASSIFICATION_COL_IDX] is None
        ]
        if not pink_rows:
            continue

        review_new_rows(pink_rows, bedrock_client, model_id, tab_name=tab_name)

        relocated = 0
        for row in pink_rows:
            target_tab = _relocation_target(tab_name, row[REVIEW_CLASSIFICATION_COL_IDX])
            if target_tab and target_tab != tab_name:
                rows.remove(row)
                tab_data.setdefault(target_tab, []).append(row)
                relocated += 1
        if relocated:
            print(f"     ↪ relocated {relocated} row(s) out of '{tab_name}' by classification")


# ============================================================================
# EXCEL OUTPUT
# ============================================================================

HEADER_FONT = Font(name="Arial", bold=True, size=10, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="003366")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_FONT = Font(name="Arial", size=10)
THIN_BORDER = Border(
    bottom=Side(style="thin", color="D0D0D0"),
)

# Data-label row fills, matching the Legend tab: Orange = new T1 data to review,
# Pink = new Non-T1 data to review. Yellow overrides both — it means the mention
# review LLM step classified the row D (No GoDaddy mention), G (Press Release),
# or H (Commercial/Ad), so it needs a human look before being kept or removed.
# Red ("Removed") is a manual designation this script never applies and never
# clears. NO_FILL is used to "age out" a row from Orange/Pink to White once it's
# no longer new.
ORANGE_FILL_ARGB = "FFFBE2D5"
PINK_FILL_ARGB = "FFF2CEEF"
RED_FILL_ARGB = "FFFF0000"
YELLOW_FILL_ARGB = "FFFFFF00"
ORANGE_FILL = PatternFill("solid", fgColor=ORANGE_FILL_ARGB)
PINK_FILL = PatternFill("solid", fgColor=PINK_FILL_ARGB)
YELLOW_FILL = PatternFill("solid", fgColor=YELLOW_FILL_ARGB)
NO_FILL = PatternFill(fill_type=None)

# Mention-review classifications that mean "not a keeper" — see REVIEW_SYSTEM_PROMPT.
YELLOW_CLASSIFICATIONS = {"D", "G", "H"}


def _row_fill(row_data: list) -> PatternFill:
    """Yellow once Review Classification lands on D/G/H (only ever set for Pink
    rows — Orange rows are never reviewed, see review_and_relocate). Otherwise
    Orange for T1 rows, Pink for Non-T1 — row.is_t1 is the single source of truth
    shared with Custom Categories (see mention_to_row), so the two never disagree."""
    if row_data[REVIEW_CLASSIFICATION_COL_IDX] in YELLOW_CLASSIFICATIONS:
        return YELLOW_FILL
    return ORANGE_FILL if getattr(row_data, "is_t1", False) else PINK_FILL


def write_tab(ws, rows: list[list]):
    """Write header + data rows to a worksheet with formatting."""
    # Header row
    for col_idx, header in enumerate(IC_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN

    # Data rows (sorted by date desc, then time desc)
    rows_sorted = sorted(rows, key=lambda r: (r[0] or datetime.min.date(), r[1] or datetime.min.time()), reverse=True)

    for row_idx, row_data in enumerate(rows_sorted, start=2):
        fill = _row_fill(row_data)
        for col_idx, value in enumerate(row_data, start=1):
            # ws.cell(..., value=value) silently *skips* setting the value when
            # value is None (its default-arg design can't tell "explicitly write
            # None" from "caller didn't pass one") — assign .value directly instead.
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.font = CELL_FONT
            cell.border = THIN_BORDER
            cell.fill = fill

    # Column widths
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    if rows:
        last_col = get_column_letter(len(IC_COLUMNS))
        ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"


# Sheet-tab colors, matching godaddy_data_june_2026.xlsx exactly (theme index + tint).
SBRL_TAB_COLOR = Color(theme=8, tint=0.7999816888943144, type="theme")
PRODUCT_TAB_COLOR = Color(theme=4, tint=0.7999816888943144, type="theme")
ANS_PRODUCT_TAB_COLOR = Color(theme=7, tint=0.7999816888943144, type="theme")  # ANS (Product) + Int. only — an outlier vs. the other Product tabs
# "Olive Green, Accent 3, Lighter 80%" — used for the Corporate group (Brand,
# Finance, Thought Leadership person tabs) and ANS Open Standard.
OLIVE_GREEN_TAB_COLOR = Color(theme=6, tint=0.7999816888943144, type="theme")

TAB_COLOR_BY_TAB = {
    "GDSBRL + Int.": SBRL_TAB_COLOR,
    "Commerce (Product) + Int.": PRODUCT_TAB_COLOR,
    "AGI (Product) + Int.": PRODUCT_TAB_COLOR,
    "Airo (Product) + Int.": PRODUCT_TAB_COLOR,
    "ANS (Product) + Int.": ANS_PRODUCT_TAB_COLOR,
    "Other (Product) + Int.": PRODUCT_TAB_COLOR,
    "ANS Open Standard": OLIVE_GREEN_TAB_COLOR,
    "Brand + Int.": OLIVE_GREEN_TAB_COLOR,
    "Finance + Int.": OLIVE_GREEN_TAB_COLOR,
    **{tab: OLIVE_GREEN_TAB_COLOR for tab in PEOPLE_TABS.values()},
}


def build_legend_sheet(wb, index: int = 0):
    """Create the Legend tab explaining the workbook's color-coding conventions."""
    ws = wb.create_sheet(title="Legend", index=index)
    ws.column_dimensions["B"].width = 21.1
    ws.column_dimensions["C"].width = 22.1

    ws["B2"] = "Data Labels"

    ws["B3"] = "Red"
    ws["B3"].fill = PatternFill("solid", fgColor="FFFF0000")
    ws["C3"] = "Removed"

    ws["B4"] = "Orange"
    ws["B4"].fill = PatternFill("solid", fgColor="FFFBE2D5")
    ws["C4"] = "Keep - T1"

    ws["B5"] = "Pink"
    ws["B5"].fill = PatternFill("solid", fgColor="FFF2CEEF")
    ws["C5"] = "Keep - Non-T1"

    ws["B6"] = "Yellow"
    ws["B6"].fill = PatternFill("solid", fgColor=YELLOW_FILL_ARGB)
    ws["C6"] = "Classified D / G / H — no GoDaddy mention, press release, or ad"

    ws["B8"] = "Tabs"

    ws["B9"] = "Pink"
    ws["B9"].fill = PatternFill("solid", fgColor="FFF2CFEF")
    ws["C9"] = "SBRL"

    ws["B10"] = "Blue"
    ws["B10"].fill = PatternFill("solid", fgColor="FFC0E5F4")
    ws["C10"] = "Product"

    ws["B11"] = "Green"
    ws["B11"].fill = PatternFill(fill_type="solid", fgColor=OLIVE_GREEN_TAB_COLOR)
    ws["C11"] = "Corporate"

    ws["B12"] = "White"
    ws["C12"] = "Source of Truth/General"

    return ws


def create_workbook(tab_data: dict[str, list[list]], output_path: str, bedrock_client=None, bedrock_model_id: str = BEDROCK_MODEL_ID):
    """Create a new IC Data workbook with all tabs. Every row is new here, so if
    bedrock_client is set, every Pink row (except NO_GODADDY_FOCUS_TABS) gets
    LLM-reviewed and possibly relocated before anything is written — Orange rows
    are left alone entirely (see review_and_relocate)."""
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    build_legend_sheet(wb)

    review_and_relocate(tab_data, bedrock_client, bedrock_model_id)

    for tab_name in TAB_ORDER:
        rows = tab_data.get(tab_name, [])
        ws = wb.create_sheet(title=tab_name[:31])  # Excel 31-char tab name limit
        if tab_name in TAB_COLOR_BY_TAB:
            ws.sheet_properties.tabColor = TAB_COLOR_BY_TAB[tab_name]
        write_tab(ws, rows)
        print(f"  📄 {tab_name}: {len(rows)} mentions")

    resort_all_tabs(wb)
    wb.save(output_path)
    print(f"\n✅ Saved to {output_path}")


def _row_sort_key(cells_info):
    """Group by data-label color (Orange -> Pink -> Yellow -> everything else),
    newest first within each group. cells_info is the list of (value, font, fill,
    border, alignment, number_format) tuples captured for one row; Date is column
    1, Time is column 2."""
    _, _, fill, *_ = cells_info[0]
    fg = fill.fgColor
    if fill.fill_type == "solid" and fg.type == "rgb" and fg.rgb == ORANGE_FILL_ARGB:
        group = 0
    elif fill.fill_type == "solid" and fg.type == "rgb" and fg.rgb == PINK_FILL_ARGB:
        group = 1
    elif fill.fill_type == "solid" and fg.type == "rgb" and fg.rgb == YELLOW_FILL_ARGB:
        group = 2
    else:
        group = 3

    date_val = cells_info[0][0]
    time_val = cells_info[1][0]
    date_ord = date_val.toordinal() if date_val else 0
    time_secs = (time_val.hour * 3600 + time_val.minute * 60 + time_val.second) if time_val else 0
    return (group, -date_ord, -time_secs)


def resort_tab(ws):
    """Re-lay-out a data tab's rows: new T1 (Orange) first, new Non-T1 (Pink)
    second, flagged new data (Yellow) third, then everything else — newest-to-oldest
    by Date/Time within each group."""
    if ws.max_row < 3:
        return  # header only (or a single data row) — nothing to reorder

    max_col = ws.max_column
    entries = [
        # cell.font/.fill/.border/.alignment return an immutable StyleProxy; copy()
        # unwraps it into a real, reassignable style object.
        [(c.value, copy(c.font), copy(c.fill), copy(c.border), copy(c.alignment), c.number_format) for c in row]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=max_col)
    ]
    entries.sort(key=_row_sort_key)

    for row_idx, cells_info in enumerate(entries, start=2):
        for col_idx, (value, font, fill, border, alignment, number_format) in enumerate(cells_info, start=1):
            # See the matching comment in write_tab: must assign .value directly,
            # not via ws.cell(..., value=value), or a None here silently leaves
            # whatever value that cell position held before the resort.
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.font = font
            cell.fill = fill
            cell.border = border
            cell.alignment = alignment
            cell.number_format = number_format


def resort_all_tabs(wb):
    """Apply resort_tab to every tab this script manages that exists in the workbook."""
    for tab_name in TAB_ORDER:
        ws_name = tab_name[:31]
        if ws_name in wb.sheetnames:
            resort_tab(wb[ws_name])


def downgrade_reviewed_rows(wb):
    """Age out last run's "new data" highlighting: any row still filled Orange or Pink
    (Keep - T1 / Keep - Non-T1) is cleared to White now that it's no longer new.
    Yellow (flagged: "GoDaddy" not found in Hit Sentence) is a standing content-quality
    flag, not a review-recency one — like Red (Removed), it's left untouched until a
    human resolves it. Only scans tabs this script manages (not foreign/manual sheets
    in the workbook)."""
    downgraded = 0
    for tab_name in TAB_ORDER:
        ws_name = tab_name[:31]
        if ws_name not in wb.sheetnames:
            continue
        ws = wb[ws_name]
        if ws.max_row < 2:
            continue
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            marker_fill = row[0].fill
            fg = marker_fill.fgColor
            if marker_fill.fill_type == "solid" and fg.type == "rgb" and fg.rgb in (ORANGE_FILL_ARGB, PINK_FILL_ARGB):
                for cell in row:
                    cell.fill = NO_FILL
                downgraded += 1
    if downgraded:
        print(f"  🎨 Downgraded {downgraded} previously-new row(s) from Orange/Pink to White")
    return downgraded


def append_to_workbook(tab_data: dict[str, list[list]], existing_path: str, output_path: str, bedrock_client=None, bedrock_model_id: str = BEDROCK_MODEL_ID):
    """Append new mentions to an existing IC Data workbook, deduplicating by Document
    ID. If bedrock_client is set, only the genuinely new rows (post-dedup) get
    LLM-reviewed and possibly relocated — never rows already in the workbook from
    a prior run."""
    wb = openpyxl.load_workbook(existing_path)

    if "Legend" not in wb.sheetnames:
        build_legend_sheet(wb)

    downgrade_reviewed_rows(wb)

    # ── Pass 1: dedup against Document IDs already saved ANYWHERE in the workbook ──
    # (not just the row's own destination tab) — review_and_relocate can move a row
    # to a different tab than the one its source stream targets, so a mention we
    # already saved under AGI last run could come back tagged for Brand this run.
    existing_ids: set = set()
    for tab_name in TAB_ORDER:
        ws_name = tab_name[:31]
        if ws_name in wb.sheetnames:
            for row in wb[ws_name].iter_rows(min_row=2, min_col=3, max_col=3, values_only=True):
                if row[0] is not None:
                    existing_ids.add(row[0])

    new_by_tab: dict[str, list] = {}
    for tab_name, candidate_rows in tab_data.items():
        if not candidate_rows:
            continue
        new_unique = [r for r in candidate_rows if r[2] not in existing_ids]
        dupes = len(candidate_rows) - len(new_unique)
        print(f"  📄 {tab_name}: {len(new_unique)} new ({dupes} dupes skipped)")
        if new_unique:
            new_by_tab[tab_name] = new_unique

    # ── Pass 2: LLM-review Pink rows, relocating A/B/C classifications ──
    review_and_relocate(new_by_tab, bedrock_client, bedrock_model_id)

    # ── Pass 3: write everything, including any tab a relocation newly touched ──
    for tab_name, new_rows in new_by_tab.items():
        if not new_rows:
            continue
        ws_name = tab_name[:31]
        if ws_name in wb.sheetnames:
            ws = wb[ws_name]
            start_row = ws.max_row + 1
            for offset, row_data in enumerate(new_rows):
                row_idx = start_row + offset
                fill = _row_fill(row_data)
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.font = CELL_FONT
                    cell.border = THIN_BORDER
                    cell.fill = fill
            print(f"  ✓ {tab_name}: wrote {len(new_rows)} row(s)")
        else:
            ws = wb.create_sheet(title=ws_name)
            if tab_name in TAB_COLOR_BY_TAB:
                ws.sheet_properties.tabColor = TAB_COLOR_BY_TAB[tab_name]
            write_tab(ws, new_rows)
            print(f"  ✓ {tab_name}: wrote {len(new_rows)} row(s) (new tab)")

    resort_all_tabs(wb)
    wb.save(output_path)
    print(f"\n✅ Saved to {output_path}")


# ============================================================================
# MAIN PIPELINE
# ============================================================================

def fetch_all_streams(client: CisionOneClient, after: str, before: str, fetch_full_text: bool = False) -> dict[str, list[list]]:
    """Fetch mentions from all 18 streams and organize by tab."""
    tab_data: dict[str, list[list]] = defaultdict(list)
    total_mentions = 0
    unmatched_person_mentions = 0
    flagged_count = 0

    for i, stream in enumerate(STREAMS, start=1):
        sid = stream["id"]
        label = stream["label"]
        tier = stream["tier"]
        tab = stream["tab"]

        print(f"\n[{i}/{len(STREAMS)}] Fetching {tier} — {label} (ID {sid}) → {tab}")

        try:
            mentions = client.get_all_mentions_chunked(sid, after, before, chunk_days=1)
        except Exception as e:
            print(f"    ❌ Error: {e}")
            continue

        if stream.get("split_by_person"):
            matched = 0
            for m in mentions:
                person_tabs = {
                    PEOPLE_LOOKUP[kw.lower()]
                    for kw in (m.get("keywords") or [])
                    if kw.lower() in PEOPLE_LOOKUP
                }
                if not person_tabs:
                    unmatched_person_mentions += 1
                    continue
                row = mention_to_row(m, stream, fetch_full_text=fetch_full_text)
                if row.hit_sentence_flagged:
                    flagged_count += 1
                for person_tab in person_tabs:
                    tab_data[person_tab].append(row)
                matched += 1
            print(f"    ✓ {len(mentions)} mentions → routed {matched} to person tabs")
        else:
            rows = [mention_to_row(m, stream, fetch_full_text=fetch_full_text) for m in mentions]
            flagged_count += sum(1 for r in rows if r.hit_sentence_flagged)
            tab_data[tab].extend(rows)
            print(f"    ✓ {len(mentions)} mentions")

        total_mentions += len(mentions)

    if unmatched_person_mentions:
        print(f"\n⚠️  {unmatched_person_mentions} Thought Leadership mention(s) didn't match any known person and were skipped")
    if flagged_count:
        suffix = " even after checking the full article" if fetch_full_text else " (run with --fetch-full-text to double-check the full article)"
        print(f"\n🟡 {flagged_count} mention(s) flagged Yellow — \"GoDaddy\" not found{suffix}")

    # Deduplicate within each tab (same article can appear in T1 and Non-T1 with different Input Names)
    # We keep both since they have different Input Names — that's intentional per the grid design
    print(f"\n📊 Total mentions across all streams: {total_mentions}")
    return dict(tab_data)


def main():
    parser = argparse.ArgumentParser(description="Generate GoDaddy IC Data grid from Cision One API")
    parser.add_argument("--after", required=True, help="Start date (YYYY-MM-DD)")
    parser.add_argument("--before", required=True, help="End date (YYYY-MM-DD)")
    parser.add_argument("--output", default=None, help="Output file path (default: GoDaddy_IC_Data_<after>_to_<before>.xlsx)")
    parser.add_argument("--append-to", default=None, help="Path to existing workbook to append to")
    parser.add_argument("--token", default=None, help="Cision One API token (or set CISION_API_TOKEN env var)")
    parser.add_argument(
        "--fetch-full-text", action="store_true",
        help="For mentions flagged Yellow (\"GoDaddy\" not in Cision's excerpt), fetch the "
             "mention's own URL and re-check the full article text before giving up. Off by "
             "default — adds one HTTP request per flagged mention.",
    )
    parser.add_argument(
        "--llm-review", action=argparse.BooleanOptionalAction, default=True,
        help="LLM-review every newly-added row for whether it's a legit GoDaddy mention, using "
             "the full article text where fetchable (Bedrock Claude). On by default — pass "
             "--no-llm-review to skip it. Requires AWS credentials with Bedrock access.",
    )
    args = parser.parse_args()

    token = args.token or os.environ.get("CISION_API_TOKEN")
    if not token:
        print("Error: No API token. Pass --token or set CISION_API_TOKEN.")
        sys.exit(1)

    after_iso = f"{args.after}T00:00:00.000Z"
    before_iso = f"{args.before}T23:59:59.000Z"

    bedrock_client = build_bedrock_client() if args.llm_review else None

    print("=" * 60)
    print("  GoDaddy IC Data Grid Generator")
    print("=" * 60)
    print(f"  Date range: {args.after} → {args.before}")
    print(f"  Streams: {len(STREAMS)}")
    print(f"  Rate limit: ~10 req/min ({REQUEST_INTERVAL}s between requests)")
    print(f"  Mode: {'Append' if args.append_to else 'New workbook'}")
    print(f"  Full-text re-check for flagged mentions: {'On' if args.fetch_full_text else 'Off'}")
    print(f"  LLM review of new mentions: {'On (' + BEDROCK_MODEL_ID + ')' if bedrock_client else 'Off'}")
    print()

    client = CisionOneClient(api_token=token)
    tab_data = fetch_all_streams(client, after_iso, before_iso, fetch_full_text=args.fetch_full_text)

    if args.append_to:
        output_path = args.output or args.append_to
        append_to_workbook(tab_data, args.append_to, output_path, bedrock_client=bedrock_client)
    else:
        output_path = args.output or f"GoDaddy_IC_Data_{args.after}_to_{args.before}.xlsx"
        create_workbook(tab_data, output_path, bedrock_client=bedrock_client)


if __name__ == "__main__":
    main()
