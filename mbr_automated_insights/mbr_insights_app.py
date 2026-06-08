import streamlit as st
import pandas as pd
import openpyxl
import json
import boto3
from io import BytesIO
from datetime import date

# ──────────────────────────────────────────────────────────────
# Page config
# ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Media Coverage Insights",
    page_icon="📊",
    layout="wide",
)

st.title("📊 Media Coverage Insights Generator")
st.markdown(
    "Upload the GoDaddy IC Data workbook and let Claude generate "
    "business-unit summaries, an executive summary, and strategic "
    "insights — all in one click."
)

st.info(
    "**How it works:** Upload the full GoDaddy IC Data workbook — red-highlighted "
    "(scrubbed) rows are automatically discarded, and you can narrow the date range "
    "to focus on a specific month or period. No manual data prep required."
)

# ──────────────────────────────────────────────────────────────
# Sheet-name → internal-key mapping  (matched by substring)
#
# Each tuple is (substring_to_match, internal_key).
# Matching is case-insensitive and checks whether the substring
# appears anywhere in the actual Excel sheet name, so small
# naming variations across yearly files are handled automatically.
# ORDER MATTERS: more-specific substrings must come before
# less-specific ones (e.g. "ans open standard" before "ans").
# ──────────────────────────────────────────────────────────────
SHEET_MAP = [
    ("gdsbrl",              "small_business_research_lab"),
    ("commerce",            "commerce"),
    ("agi",                 "agi"),
    ("airo",                "airo"),
    ("ans open standard",   "ans_open_standard"),
    ("ans",                 "ans"),
    ("other",               "other"),
    ("brand identity",      "_skip_brand_identity"),
    ("brand",               "brand"),
    ("finance + int",       "finance"),
    ("finance",             "finance"),
    ("aman bhutani",        "aman_bhutani"),
    ("gourav pani",         "gourav_pani"),
    ("kasturi mudulodu",    "kasturi_mudulodu"),
    ("mark mccaffrey",      "mark_mccaffrey"),
    ("jared sine",          "jared_sine"),
    ("travis muhlestein",   "travis_muhlestein"),
    ("demetria",            "demetria"),
    ("berea schaffer",      "berea_schaffer"),
    ("general",             "general"),
]

COLUMNS_TO_KEEP = ["Date", "Title", "Hit Sentence"]

# ──────────────────────────────────────────────────────────────
# Business-unit definitions
#
# Each entry: (internal_key, display_label, list_of_source_keys)
# If source_keys is None the key is read directly from the
# parsed sheets dict; otherwise sources are concatenated.
# ──────────────────────────────────────────────────────────────
BUSINESS_UNITS = [
    ("small_business_research_lab", "🔬 Small Business Research Lab", None),
    ("product", "🏭 Product", ["commerce", "agi", "airo", "ans", "other"]),
    ("ans_open_standard", "🌐 ANS Open Standard", None),
    ("brand", "🎨 Brand", None),
    (
        "thought_leadership",
        "👔 Thought Leadership",
        [
            "aman_bhutani",
            "gourav_pani",
            "kasturi_mudulodu",
            "mark_mccaffrey",
            "jared_sine",
            "travis_muhlestein",
            "demetria",
            "berea_schaffer",
        ],
    ),
    ("financial", "💰 Financial", ["finance"]),
    (
        "corporate",
        "🏢 Corporate",
        [
            "brand",
            "aman_bhutani",
            "gourav_pani",
            "kasturi_mudulodu",
            "mark_mccaffrey",
            "jared_sine",
            "travis_muhlestein",
            "demetria",
            "berea_schaffer",
            "finance",
        ],
    ),
]

# Red-fill color code used in the IC Data grid to mark scrubbed rows
_RED_RGB = "FFFF0000"


# ──────────────────────────────────────────────────────────────
# Data helpers
# ──────────────────────────────────────────────────────────────
def _match_sheet(sheet_name: str) -> str | None:
    """Return the internal key for a given Excel sheet name, or None."""
    lower = sheet_name.lower()
    for substring, key in SHEET_MAP:
        if substring in lower:
            return key
    return None


def _is_red_fill(fill) -> bool:
    """Check whether an openpyxl cell fill is red (FFFF0000, solid)."""
    if fill.patternType != "solid":
        return False
    fg = fill.fgColor
    if fg is None or fg.type != "rgb":
        return False
    try:
        return fg.rgb == _RED_RGB
    except Exception:
        return False


@st.cache_data(show_spinner="Loading workbook and filtering red-highlighted rows …")
def load_and_process(file_bytes: bytes) -> tuple[dict[str, pd.DataFrame], dict[str, int]]:
    """Single-pass load: extract data via openpyxl while auto-discarding
    rows whose Date cell is highlighted red.

    Returns
    -------
    datasets : dict mapping unit keys → DataFrames (columns: Date, Title, Hit Sentence)
    red_counts : dict mapping sheet internal keys → number of red rows discarded
    """
    buf = BytesIO(file_bytes)
    wb = openpyxl.load_workbook(buf, data_only=True)

    raw: dict[str, pd.DataFrame] = {}
    red_counts: dict[str, int] = {}
    matched_keys: set[str] = set()

    for sheet_name in wb.sheetnames:
        key = _match_sheet(sheet_name)
        if key is None or key.startswith("_skip") or key in matched_keys:
            continue
        matched_keys.add(key)

        ws = wb[sheet_name]

        # -- Read header row --------------------------------------------------
        header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=False), None)
        if header_cells is None:
            raw[key] = pd.DataFrame(columns=COLUMNS_TO_KEEP)
            red_counts[key] = 0
            continue

        headers = [c.value for c in header_cells]

        # Locate the Date column for red-fill detection
        date_col_idx: int | None = None
        for i, h in enumerate(headers):
            if h == "Date":
                date_col_idx = i
                break

        # -- Single pass: collect data rows, skip red-highlighted ones ---------
        kept_rows: list[list] = []
        red = 0
        for row in ws.iter_rows(min_row=2, values_only=False):
            if date_col_idx is not None:
                date_cell = row[date_col_idx]
                if _is_red_fill(date_cell.fill):
                    red += 1
                    continue
            kept_rows.append([c.value for c in row])

        df = pd.DataFrame(kept_rows, columns=headers) if kept_rows else pd.DataFrame(columns=headers)
        if "Date" in df.columns:
            # Some cells store dates as Excel serial numbers (ints)
            # rather than datetime objects — normalize both forms.
            _EXCEL_EPOCH = pd.Timestamp("1899-12-30")
            def _to_dt(val):
                if isinstance(val, (int, float)) and val > 0:
                    try:
                        return _EXCEL_EPOCH + pd.Timedelta(days=int(val))
                    except Exception:
                        return pd.NaT
                return val
            df["Date"] = df["Date"].apply(_to_dt)
            df["Date"] = pd.to_datetime(df["Date"], errors="coerce")

        raw[key] = df
        red_counts[key] = red

    wb.close()

    # ── Warn about missing sheets ──
    expected = {
        "small_business_research_lab", "commerce", "agi", "airo", "ans",
        "ans_open_standard", "other", "brand", "finance",
        "aman_bhutani", "gourav_pani", "kasturi_mudulodu",
        "mark_mccaffrey", "jared_sine",
        "travis_muhlestein", "demetria", "berea_schaffer",
    }
    missing = expected - set(raw.keys())
    if missing:
        st.warning(
            f"The following expected sheets were not matched: "
            f"{', '.join(sorted(missing))}. They will be treated as empty."
        )

    # ── Helper: select analysis columns ──
    def cols(k: str) -> pd.DataFrame:
        if k not in raw:
            return pd.DataFrame(columns=COLUMNS_TO_KEEP)
        df = raw[k]
        return df[[c for c in COLUMNS_TO_KEEP if c in df.columns]]

    # ── Build combined datasets ──
    datasets: dict[str, pd.DataFrame] = {}
    for unit_key, _label, sources in BUSINESS_UNITS:
        if sources is None:
            datasets[unit_key] = cols(unit_key)
        else:
            datasets[unit_key] = pd.concat(
                [cols(s) for s in sources], ignore_index=True
            )

    datasets["all_coverage"] = (
        pd.concat(
            [cols(k) for k in raw if not k.startswith("_skip") and k != "general"],
            ignore_index=True,
        )
        .drop_duplicates()
    )

    return datasets, red_counts


def apply_date_filter(
    datasets: dict[str, pd.DataFrame],
    start: date,
    end: date,
) -> dict[str, pd.DataFrame]:
    """Return a shallow copy of datasets with rows filtered to [start, end]."""
    s = pd.Timestamp(start)
    e = pd.Timestamp(end) + pd.Timedelta(days=1) - pd.Timedelta(seconds=1)
    filtered: dict[str, pd.DataFrame] = {}
    for key, df in datasets.items():
        if "Date" in df.columns and not df.empty:
            mask = df["Date"].between(s, e)
            filtered[key] = df.loc[mask].reset_index(drop=True)
        else:
            filtered[key] = df
    return filtered


# ──────────────────────────────────────────────────────────────
# AWS Bedrock configuration
# ──────────────────────────────────────────────────────────────
AWS_REGION = st.secrets["BEDROCK_REGION"]
AWS_ACCESS_KEY_ID = st.secrets["AWS_ACCESS_KEY_ID"]
AWS_SECRET_ACCESS_KEY = st.secrets["AWS_SECRET_ACCESS_KEY"]
CLAUDE_MODEL_ID = st.secrets["BEDROCK_MODEL_ID"]


# ──────────────────────────────────────────────────────────────
# LLM helpers
# ──────────────────────────────────────────────────────────────
def call_claude(bedrock_client, prompt: str, max_tokens: int = 400) -> str:
    """Invoke Claude via AWS Bedrock."""
    body = json.dumps({
        "anthropic_version": "bedrock-2023-05-31",
        "max_tokens": max_tokens,
        "messages": [{"role": "user", "content": prompt}],
        "temperature": 0.5,
    })
    response = bedrock_client.invoke_model(
        modelId=CLAUDE_MODEL_ID,
        contentType="application/json",
        accept="application/json",
        body=body,
    )
    result = json.loads(response["body"].read())
    return result["content"][0]["text"]


def generate_coverage_summary(client, df: pd.DataFrame, unit_name: str) -> str:
    """Generate a BU-level coverage summary."""
    coverage_data = df[["Date", "Title", "Hit Sentence"]].to_string(index=False)
    min_date, max_date = df["Date"].min(), df["Date"].max()
    date_range = (
        f"{min_date.strftime('%Y-%m-%d')} to {max_date.strftime('%Y-%m-%d')}"
        if pd.notna(min_date) else "N/A"
    )
    prompt = f"""You are analyzing media coverage data for the {unit_name} area of GoDaddy.

Dataset: {unit_name}
Total articles: {len(df)}
Date range: {date_range}

Here is the complete coverage data:
{coverage_data}

Please provide a brief, concise summary (3-5 bullet points, ~100-150 words) that describes:
- The main themes and topics covered
- Any notable trends or patterns
- Specific coverage examples generating and driving these insights

GUIDELINES:
- Keep the summary factual and focused on what the data shows.
- Do not make up ANY information, numbers or figures.
- Use only what is in the provided data to generate insights."""
    return call_claude(client, prompt, max_tokens=300)


def generate_executive_summary(client, summaries: dict[str, str]) -> str:
    """Synthesize individual BU summaries into an executive overview."""
    summaries_text = "\n\n".join(
        f"{name.replace('_', ' ').upper()}:\n{summary}"
        for name, summary in summaries.items()
    )
    prompt = f"""You are creating an executive summary of media coverage of GoDaddy across all business units and coverage areas.

Below are summaries of coverage from each area:

{summaries_text}

Please create a concise executive summary (4-6 bullet points, ~150-200 words) that:
- Synthesizes the key themes across all coverage areas
- Highlights the most significant trends or patterns
- Provides a holistic view of the organization's media presence
- Notes any notable differences or commonalities between the coverage areas
- Provides specific coverage pieces generating and driving these insights

Keep the summary factual and focused on what the data showed."""
    return call_claude(client, prompt, max_tokens=400)


def generate_overall_insights(client, summaries: dict[str, str]) -> str:
    """Generate high-level strategic insights from the full set of section summaries."""
    summaries_text = "\n\n".join(
        f"{name.replace('_', ' ').upper()}:\n{summary}"
        for name, summary in summaries.items()
    )
    prompt = f"""You are a senior media strategist reviewing ALL coverage insights generated for GoDaddy across every business unit and coverage area.

Below are the individual section summaries that were produced:

{summaries_text}

Please generate THREE to FIVE high-level strategic insights that synthesize
patterns across the entire grid. Format your response as bullet points
starting with asterisks (*).

Each insight should:
- Be one concise sentence or two short sentences
- Surface cross-cutting patterns, trends, or narrative arcs that span
  multiple coverage areas
- Reference specific themes, product launches, campaigns, or executive
  visibility where relevant
- Call out opportunities or risks visible only at the aggregate level
  (e.g. narrative concentration, coverage gaps, message alignment)

Do NOT simply repeat individual section summaries — your value is in
connecting the dots between them."""
    return call_claude(client, prompt, max_tokens=500)


def enrich_insights_with_examples(
    client, insights_text: str, all_coverage_df: pd.DataFrame
) -> str:
    """Match each overall insight to 3-5 specific coverage articles."""
    # Build a compact, deduplicated article list (Date + Title)
    articles = (
        all_coverage_df[["Date", "Title"]]
        .dropna(subset=["Title"])
        .drop_duplicates(subset="Title")
        .sort_values("Date")
    )

    # Safety cap — if the dataset is very large, sample to keep the
    # prompt within comfortable context-window limits.
    MAX_ARTICLES = 3000
    if len(articles) > MAX_ARTICLES:
        articles = articles.sample(MAX_ARTICLES, random_state=42).sort_values("Date")

    coverage_list = "\n".join(
        f"[{row.Date.strftime('%Y-%m-%d')}] {row.Title}"
        for row in articles.itertuples()
        if pd.notna(row.Date)
    )

    prompt = f"""Below are strategic insights about GoDaddy's media coverage, followed by
a list of all coverage articles (date + title) from the analysis period.

INSIGHTS:
{insights_text}

COVERAGE ARTICLES:
{coverage_list}

For EACH insight above, select 3 to 5 specific articles from the COVERAGE
ARTICLES list that best support or illustrate that insight.

Format your response by repeating each insight as a bullet point (*),
followed by the supporting articles as indented sub-bullets (  —), each
with the date and exact title from the list.

Example format:
* [Insight text here]
  — [2026-01-15] Article Title Here
  — [2026-02-03] Another Article Title
  — [2026-03-20] Third Supporting Article

RULES:
- Use the EXACT titles from the COVERAGE ARTICLES list. Do not paraphrase
  or invent titles.
- Choose articles that genuinely demonstrate the insight, not just ones
  that share a keyword.
- Prefer articles spread across different dates when possible to show the
  trend or pattern described in the insight."""

    return call_claude(client, prompt, max_tokens=1500)


# ──────────────────────────────────────────────────────────────
# Sidebar
# ──────────────────────────────────────────────────────────────
with st.sidebar:
    st.header("ℹ️ Reference")
    st.markdown("**Insight sections generated:**")
    for _key, label, _src in BUSINESS_UNITS:
        st.caption(label)
    st.divider()
    st.caption(
        "Sheets are matched by name (case-insensitive substring), "
        "so column order or extra tabs won't break parsing."
    )
    st.caption(
        "Rows with a red-highlighted Date cell are automatically "
        "discarded before analysis."
    )

# ──────────────────────────────────────────────────────────────
# Main area — file upload → filter → run
# ──────────────────────────────────────────────────────────────
uploaded_file = st.file_uploader(
    "Upload the GoDaddy IC Data workbook (.xlsx)", type=["xlsx", "xls"]
)

if uploaded_file:
    datasets_raw, red_counts = load_and_process(uploaded_file.read())

    # ── Red-row discard summary ──
    total_red = sum(red_counts.values())
    if total_red > 0:
        sheets_hit = {k: v for k, v in red_counts.items() if v > 0}
        detail = ", ".join(
            f"{k.replace('_', ' ').title()} ({v:,})" for k, v in sheets_hit.items()
        )
        st.success(
            f"🟥 **Auto-discarded {total_red:,} red-highlighted rows** across "
            f"{len(sheets_hit)} sheet(s): {detail}"
        )

    # ── Detect date bounds from all_coverage ──
    all_dates = datasets_raw["all_coverage"]["Date"].dropna()
    if all_dates.empty:
        st.error("No valid dates found in the uploaded workbook.")
        st.stop()

    data_min = all_dates.min().date()
    data_max = all_dates.max().date()

    # ── Date range filter ──
    st.subheader("📅 Date Range Filter")
    filter_cols = st.columns([1, 1, 2])
    with filter_cols[0]:
        start_date = st.date_input(
            "Start date",
            value=data_min,
            min_value=data_min,
            max_value=data_max,
            key="start_date",
        )
    with filter_cols[1]:
        end_date = st.date_input(
            "End date",
            value=data_max,
            min_value=data_min,
            max_value=data_max,
            key="end_date",
        )
    with filter_cols[2]:
        st.caption(
            f"Data spans **{data_min.strftime('%b %d, %Y')}** to "
            f"**{data_max.strftime('%b %d, %Y')}**. Narrow the window "
            "to focus insights on a specific period."
        )

    if start_date > end_date:
        st.error("Start date must be on or before end date.")
        st.stop()

    # Apply filter
    datasets = apply_date_filter(datasets_raw, start_date, end_date)

    # ── Dataset overview (filtered) ──
    st.subheader("📋 Dataset Overview")
    date_range_active = start_date != data_min or end_date != data_max
    if date_range_active:
        st.caption(
            f"Showing rows from **{start_date.strftime('%b %d, %Y')}** to "
            f"**{end_date.strftime('%b %d, %Y')}**"
        )

    display_units = [(k, l) for k, l, _ in BUSINESS_UNITS]
    stat_cols = st.columns(len(display_units))
    for col, (key, label) in zip(stat_cols, display_units):
        n = len(datasets.get(key, pd.DataFrame()))
        col.metric(label.split(" ", 1)[1], f"{n:,} rows")

    st.divider()

    if st.button("🚀 Generate Insights", type="primary", use_container_width=True):
        # ── Bedrock client ──
        try:
            bedrock_runtime = boto3.client(
                "bedrock-runtime",
                region_name=AWS_REGION,
                aws_access_key_id=AWS_ACCESS_KEY_ID,
                aws_secret_access_key=AWS_SECRET_ACCESS_KEY,
            )
        except Exception as e:
            st.error(f"Failed to connect to AWS Bedrock: {e}")
            st.stop()

        summaries: dict[str, str] = {}
        total_steps = len(BUSINESS_UNITS) + 3  # +1 exec, +1 insights, +1 enrichment

        # ── Business-unit summaries ──
        st.subheader("📊 Coverage Area Summaries")
        progress = st.progress(0, text="Starting analysis …")

        for idx, (unit_key, unit_label, _sources) in enumerate(BUSINESS_UNITS):
            progress.progress(
                idx / total_steps,
                text=f"Summarizing {unit_label} …",
            )
            df = datasets[unit_key]
            if df.empty:
                summaries[unit_key] = "_No coverage data for this period._"
                with st.expander(unit_label, expanded=False):
                    st.markdown("_No coverage data for this period._")
                continue

            summary = generate_coverage_summary(
                bedrock_runtime,
                df,
                unit_label.split(" ", 1)[1],  # drop emoji prefix
            )
            summaries[unit_key] = summary
            with st.expander(unit_label, expanded=False):
                st.markdown(summary)

        # ── Executive summary ──
        progress.progress(
            len(BUSINESS_UNITS) / total_steps,
            text="Generating executive summary …",
        )
        st.subheader("📋 Executive Summary")
        exec_summary = generate_executive_summary(bedrock_runtime, summaries)
        st.markdown(exec_summary)

        # ── Overall insights (synthesized from section summaries) ──
        progress.progress(
            (len(BUSINESS_UNITS) + 1) / total_steps,
            text="Generating overall insights …",
        )
        insights_raw = generate_overall_insights(bedrock_runtime, summaries)

        # ── Enrich insights with supporting coverage examples ──
        progress.progress(
            (len(BUSINESS_UNITS) + 2) / total_steps,
            text="Matching insights to coverage examples …",
        )
        st.subheader("💡 Overall Insights")
        insights_enriched = enrich_insights_with_examples(
            bedrock_runtime, insights_raw, datasets["all_coverage"]
        )
        st.markdown(insights_enriched)

        progress.progress(1.0, text="✅ Analysis complete!")

else:
    st.info("Upload the GoDaddy IC Data workbook to get started.")
